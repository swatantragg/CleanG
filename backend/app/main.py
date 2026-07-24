import os
from contextlib import asynccontextmanager

from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from slowapi.errors import RateLimitExceeded
from sqlalchemy import select, text
from sqlalchemy.exc import DBAPIError, OperationalError, SQLAlchemyError
from starlette.middleware.trustedhost import TrustedHostMiddleware
from starlette.responses import JSONResponse, Response

from .config import get_settings
from .core.csrf import csrf_protect
from .core.dynamic_columns import make_attr, quote_ident, sync_custom_columns
from .core.limiter import limiter
from .core.scheduler import shutdown_scheduler, start_scheduler
from .database import Base, SessionLocal, engine
from .models import MasterColumn, User, UserRole
from .routers import (
    auth,
    branches,
    clean,
    files,
    master,
    mlc,
    prs,
    reports,
    standardize,
    users,
)
from .security import hash_password

settings = get_settings()

MASTER_FILE = os.path.join(os.path.dirname(__file__), "data", "master_output_format.xlsx")


def _seed_master_columns(db) -> None:
    """Load the canonical output schema from the bundled master workbook."""
    if db.scalar(select(MasterColumn).limit(1)) is not None:
        return
    wb = load_workbook(MASTER_FILE, data_only=True)
    ws = wb.active
    position = 0
    for cell in ws[1]:
        value = cell.value
        if value is None or str(value).strip() == "":
            continue
        position += 1
        db.add(MasterColumn(position=position, name=str(value).strip()))
    db.commit()


def _migrate(db) -> None:
    """Add the review-overlay columns to pre-existing tables (idempotent).

    `create_all` only creates missing tables, never alters existing ones, so the
    `corrections`/`dropped` columns are added here for databases created before
    cleaning moved fully in-memory.
    """
    db.execute(text(
        "ALTER TABLE uploaded_files "
        "ADD COLUMN IF NOT EXISTS corrections JSONB NOT NULL DEFAULT '{}'::jsonb"
    ))
    db.execute(text(
        "ALTER TABLE uploaded_files "
        "ADD COLUMN IF NOT EXISTS dropped JSONB NOT NULL DEFAULT '[]'::jsonb"
    ))
    db.execute(text(
        "ALTER TABLE uploaded_files "
        "ADD COLUMN IF NOT EXISTS accepted JSONB NOT NULL DEFAULT '[]'::jsonb"
    ))
    db.execute(text(
        "ALTER TABLE uploaded_files "
        "ADD COLUMN IF NOT EXISTS constants JSONB NOT NULL DEFAULT '{}'::jsonb"
    ))
    # Merge-origin marks for review cells (tagged "Merged value" vs a hand edit).
    db.execute(text(
        "ALTER TABLE uploaded_files "
        "ADD COLUMN IF NOT EXISTS merged_cells JSONB NOT NULL DEFAULT '{}'::jsonb"
    ))
    # Forced-password-rotation flag, added for databases created before it existed.
    db.execute(text(
        "ALTER TABLE users "
        "ADD COLUMN IF NOT EXISTS must_change_password BOOLEAN NOT NULL DEFAULT FALSE"
    ))
    # Custom (user-added) master columns: a flag + physical column name on the
    # schema table. Their values are now REAL columns on master_data.
    db.execute(text(
        "ALTER TABLE master_columns "
        "ADD COLUMN IF NOT EXISTS custom BOOLEAN NOT NULL DEFAULT FALSE"
    ))
    db.execute(text(
        "ALTER TABLE master_columns ADD COLUMN IF NOT EXISTS attr VARCHAR"
    ))
    db.commit()
    # Promote any legacy `extras` JSON values into real columns, then retire the
    # bag entirely (idempotent: a no-op on databases without the legacy column).
    _migrate_custom_columns(db)
    db.execute(text("ALTER TABLE master_data DROP COLUMN IF EXISTS extras"))
    # Cleaned rows are no longer persisted — drop the legacy table so its stale
    # rows can't block file/branch deletion via the old foreign key.
    db.execute(text("DROP TABLE IF EXISTS cleaned_rows"))
    db.commit()


def _migrate_custom_columns(db) -> None:
    """Back-fill custom columns from the legacy `master_data.extras` JSON bag into
    dedicated real columns. Idempotent and safe on fresh databases.

    For each custom master column: ensure it has a physical `attr`, add the real
    column (metadata-only with its default), and copy any value that lived in the
    `extras` bag into it. Identifiers are regex-validated and quoted.
    """
    has_extras = db.scalar(text(
        "SELECT 1 FROM information_schema.columns "
        "WHERE table_name = 'master_data' AND column_name = 'extras'"
    ))
    rows = db.execute(text(
        "SELECT id, name, attr FROM master_columns WHERE custom = TRUE"
    )).all()
    taken = {r.attr for r in rows if r.attr}
    for r in rows:
        attr = r.attr
        if not attr:
            attr = make_attr(r.name, taken)
            taken.add(attr)
            db.execute(
                text("UPDATE master_columns SET attr = :a WHERE id = :i"),
                {"a": attr, "i": r.id},
            )
        db.execute(text(
            f"ALTER TABLE master_data ADD COLUMN IF NOT EXISTS "
            f"{quote_ident(attr)} VARCHAR NOT NULL DEFAULT ''"
        ))
        if has_extras:
            db.execute(
                text(
                    f"UPDATE master_data SET {quote_ident(attr)} = extras ->> :n "
                    "WHERE extras ->> :n IS NOT NULL"
                ),
                {"n": r.name},
            )
    db.commit()


def init_db() -> None:
    """Create tables and seed the bootstrap admin + master schema if missing."""
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        _migrate(db)
        has_user = db.scalar(select(User).limit(1))
        if has_user is None:
            db.add(
                User(
                    email=settings.admin_email,
                    full_name=settings.admin_name,
                    hashed_password=hash_password(settings.admin_password),
                    role=UserRole.admin,
                    # The bootstrap password comes from the environment and is
                    # known to whoever deployed it — force a change on first login.
                    must_change_password=True,
                )
            )
            db.commit()
        _seed_master_columns(db)
        # Attach every custom column to the ORM mapper so this process can
        # read/write them like built-ins from the first request.
        sync_custom_columns(db)
    finally:
        db.close()


class _DBState:
    """Whether the database is currently usable, for the readiness gate below."""

    ready = False
    error = ""


db_state = _DBState()

# Shown to the user (login page included) whenever the database is unreachable.
DB_DOWN_DETAIL = (
    "The database is temporarily unreachable, so we can't sign you in right now. "
    "The service reconnects automatically — please retry in a minute."
)


def mark_db_down(exc: Exception) -> None:
    """Flag the database as unusable so the supervisor re-initialises on recovery."""
    db_state.ready = False
    db_state.error = type(exc).__name__


async def _db_supervisor() -> None:
    """Keep trying to initialise the database, forever, in the background.

    Startup must never *block* on the database. It used to: `init_db()` ran inside
    the lifespan, so while the (remote, shared) Postgres was down the app never
    finished starting — uvicorn sat at "Waiting for application startup", nginx and
    Traefik had nothing to route to, and every request, including the login page's,
    died as an opaque 502/503. Restarting the process could not fix that, because
    the database, not the app, was the thing that was down.

    So the app now starts immediately and this task owns the database: it retries
    with backoff while the database is down, marks it ready once `init_db()`
    succeeds, and picks the work back up if a later request finds the connection
    dead (see `mark_db_down`). Meanwhile requests get a clear 503 instead of a 502,
    and the moment the database comes back the app heals with no restart.
    """
    import asyncio
    import logging

    log = logging.getLogger("uvicorn.error")
    max_delay = float(os.getenv("DB_RETRY_MAX_DELAY", "30"))
    delay = 2.0
    scheduler_started = False

    while True:
        if not db_state.ready:
            try:
                await asyncio.to_thread(init_db)
                db_state.ready = True
                db_state.error = ""
                delay = 2.0
                log.info("Database ready; API fully operational.")
                if not scheduler_started:
                    start_scheduler()  # daily report email (10:30 IST by default)
                    scheduler_started = True
            except Exception as exc:  # noqa: BLE001 — any failure must keep retrying
                db_state.error = type(exc).__name__
                log.warning(
                    "Database unavailable (%s); serving 503 on data routes and "
                    "retrying in %.0fs.", type(exc).__name__, delay
                )
                await asyncio.sleep(delay)
                delay = min(delay * 2, max_delay)
                continue
        await asyncio.sleep(5)


@asynccontextmanager
async def lifespan(app: FastAPI):
    import asyncio

    supervisor = asyncio.create_task(_db_supervisor())
    try:
        yield
    finally:
        supervisor.cancel()
        shutdown_scheduler()


def _rate_limit_handler(request: Request, exc: RateLimitExceeded) -> Response:
    return Response(
        '{"detail":"Too many requests. Please slow down and try again."}',
        status_code=429,
        media_type="application/json",
    )


app = FastAPI(title="MRM Cleanser API", version="0.1.0", lifespan=lifespan)

# Rate limiting (slowapi): the limiter is shared via app.core.limiter so routers
# can decorate individual endpoints (e.g. login) with @limiter.limit(...).
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_handler)

# Reject requests with an unexpected Host header (host-header poisoning) unless
# the allowlist is left as the "*" wildcard (local dev / not yet configured).
if "*" not in settings.trusted_host_list:
    app.add_middleware(
        TrustedHostMiddleware, allowed_hosts=settings.trusted_host_list
    )

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origin_list,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    # The CSRF token travels in this header, so it must be allowed cross-origin.
    allow_headers=["Authorization", "Content-Type", settings.csrf_header_name],
)

# CSRF double-submit enforcement for cookie-authenticated, state-changing calls.
app.middleware("http")(csrf_protect)


_SECURITY_HEADERS = {
    "X-Content-Type-Options": "nosniff",
    "X-Frame-Options": "DENY",
    "Referrer-Policy": "no-referrer",
    "Cross-Origin-Opener-Policy": "same-origin",
    # API responses never need to be embedded or to load remote resources.
    "Content-Security-Policy": "default-src 'none'; frame-ancestors 'none'",
    # Ignored over plain HTTP; enforced once the app is served behind TLS.
    "Strict-Transport-Security": "max-age=31536000; includeSubDomains",
}


@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    for key, value in _SECURITY_HEADERS.items():
        response.headers.setdefault(key, value)
    return response


# Paths that must answer even while the database is down: the health probe (the
# container is alive and must keep receiving traffic so users get the message
# below instead of the proxy's bare 502) and the static SPA served by nginx.
_DB_FREE_PATHS = frozenset({"/api/health"})


@app.middleware("http")
async def require_database(request: Request, call_next):
    """Answer data routes with a clear 503 while the database is unreachable."""
    path = request.url.path
    if path.startswith("/api/") and path not in _DB_FREE_PATHS and not db_state.ready:
        return JSONResponse(
            {"detail": DB_DOWN_DETAIL}, status_code=503, headers={"Retry-After": "30"}
        )
    return await call_next(request)


@app.exception_handler(SQLAlchemyError)
async def database_error(request: Request, exc: SQLAlchemyError) -> Response:
    """A database that dies *after* startup gets the same treatment as one that
    was never up: flag it so the supervisor re-initialises on recovery, and tell
    the caller what happened instead of leaking a 500."""
    if isinstance(exc, (OperationalError, DBAPIError)):
        mark_db_down(exc)
        return JSONResponse(
            {"detail": DB_DOWN_DETAIL}, status_code=503, headers={"Retry-After": "30"}
        )
    raise exc


app.include_router(auth.router)
app.include_router(users.router)
app.include_router(branches.router)
app.include_router(master.router)
app.include_router(files.router)
app.include_router(clean.router)
app.include_router(standardize.router)
app.include_router(prs.router)
app.include_router(mlc.router)
app.include_router(reports.router)


@app.get("/api/health")
def health():
    """Liveness, not readiness. This stays 200 with the database down on purpose:
    it reports the process is alive so the orchestrator keeps routing traffic here
    (killing/restarting the container cannot fix a database that is down, and
    pulling it out of the load balancer only replaces our explanatory 503 with the
    proxy's bare 502). `database` carries the real state."""
    return {
        "status": "ok" if db_state.ready else "degraded",
        "database": "up" if db_state.ready else "down",
        "detail": "" if db_state.ready else db_state.error,
    }
