"""Cleaning pipeline.

Given a branch's source files (CSV or XLSX), a primary key, and an output-column
selection (from a preset or a custom pick), the pipeline parses every file, merges
rows across files on the primary key, and writes ONE cleaned CSV (kind='cleaned').

Column names are matched case-insensitively (e.g. "ISRC" == "isrc"), so files that
label the same field differently still line up.
"""
from __future__ import annotations

import asyncio
import csv
import io
import os
import re
import unicodedata
from collections import Counter
from difflib import SequenceMatcher

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from . import settings
from .config_data import FIELD_MAP
from .models import Branch, File
from .storage import StorageBackend

try:  # mojibake repair (e.g. "à¨¸à©" → real characters); graceful if missing.
    from ftfy import fix_text as _fix_text
except Exception:  # pragma: no cover
    def _fix_text(s: str) -> str:
        return s


def _norm(s) -> str:
    return " ".join(str(s).split()).lower()


# Normalized alias table: many source/preset column names → one canonical field key
# (e.g. "Artist Name", "Singer", "Artist" all → "singer"). Lets preset output columns
# pull values from differently-named source columns.
_FIELD_CANON = {_norm(k): v for k, v in FIELD_MAP.items()}


def _canon(name: str) -> str:
    """Canonical field key for a column name; falls back to the name itself."""
    n = _norm(name)
    return _FIELD_CANON.get(n, n)


# ---- Step 2: clean & standardize individual cell values ----

_MONTHS = {m: i + 1 for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
)}
_DURATION_COL = re.compile(r"\b(duration|run ?time|length)\b", re.I)


def _clean_text(v) -> str:
    """Repair mojibake/encoding, then trim and collapse internal whitespace."""
    if v is None:
        return ""
    s = _fix_text(str(v))
    return " ".join(s.split())


def _norm_date(s: str) -> str:
    """Standardize common date formats → YYYY-MM-DD. Unparseable input is returned as-is."""
    s = s.strip()
    if not s or re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s

    # numeric: d/m/y or m/d/y (also '.' or '-' separators)
    m = re.fullmatch(r"(\d{1,4})[/.\-](\d{1,4})[/.\-](\d{1,4})", s)
    if m:
        raw = list(m.groups())
        nums = [int(x) for x in raw]
        yi = next((i for i, t in enumerate(raw) if len(t) == 4), None)
        if yi is None:
            yi = next((i for i, v in enumerate(nums) if v > 31), 2)
        year = nums[yi] + (2000 if nums[yi] < 100 else 0)
        (i1, v1), (i2, v2) = [(i, nums[i]) for i in range(3) if i != yi]
        if v1 <= 12 < v2:
            day, month = v2, v1
        elif v2 <= 12 < v1:
            day, month = v1, v2
        else:
            day, month = v1, v2  # ambiguous → day-first
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f"{year:04d}-{month:02d}-{day:02d}"
        return s

    # textual month: "9-Jan-26" / "9 Jan 2026", or "Jan 9, 2026"
    day = mon = yr = None
    m = re.fullmatch(r"(\d{1,2})[ \-]([A-Za-z]{3,9})[ \-,]+(\d{2,4})", s)
    if m:
        day, mon, yr = m.group(1), m.group(2), m.group(3)
    else:
        m = re.fullmatch(r"([A-Za-z]{3,9})[ \-]+(\d{1,2})[ \-,]+(\d{2,4})", s)
        if m:
            mon, day, yr = m.group(1), m.group(2), m.group(3)
    if mon and mon.lower()[:3] in _MONTHS:
        year = int(yr) + (2000 if int(yr) < 100 else 0)
        return f"{year:04d}-{_MONTHS[mon.lower()[:3]]:02d}-{int(day):02d}"
    return s


def _norm_duration(s: str) -> str:
    """Standardize song length → MM:SS (or HH:MM:SS). '3m 45s', '225 sec', '03:45' all → '03:45'."""
    s = s.strip().lower()
    if not s:
        return s
    m = re.fullmatch(r"(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?", s)
    if m:
        parts = [int(x) for x in m.groups() if x is not None]
        if len(parts) == 2:
            return f"{parts[0]:02d}:{parts[1]:02d}"
        return f"{parts[0]:02d}:{parts[1]:02d}:{parts[2]:02d}"
    m = re.fullmatch(r"(?:(\d+)\s*(?:m|min|mins|minutes?))?\s*(?:(\d+)\s*(?:s|sec|secs|seconds?))?", s)
    if m and (m.group(1) or m.group(2)):
        total = int(m.group(1) or 0) * 60 + int(m.group(2) or 0)
        return f"{total // 60:02d}:{total % 60:02d}"
    if re.fullmatch(r"\d+", s):  # bare seconds, e.g. "225"
        total = int(s)
        return f"{total // 60:02d}:{total % 60:02d}"
    return s


def _parse(filename: str | None, data: bytes) -> tuple[list[str], list[list[str]]]:
    """Return (headers, rows) for a CSV or XLSX file. rows are aligned to headers."""
    ext = os.path.splitext(filename or "")[1].lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            wb.close()
            return [], []
        headers = [("" if c is None else str(c)).strip() for c in header]
        body = [["" if c is None else str(c) for c in row] for row in rows_iter]
        wb.close()
        return headers, body

    # CSV (and anything else) — decode and read. newline="" lets csv handle CRLF /
    # quoted multi-line fields itself (Excel exports use \r\n).
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text, newline=""))
    all_rows = [r for r in reader]
    if not all_rows:
        return [], []
    headers = [c.strip() for c in all_rows[0]]
    return headers, all_rows[1:]


# ---- Build the cleaned, primary-key-deduplicated master ----
#
# Cleansing model (generalised — identical for every preset and for custom column picks):
#   1. DEDUP strictly on the user-selected primary key (a hashmap, O(n)). Rows sharing a
#      key are the same record → merged; different keys → kept as distinct records. No
#      fuzzy cross-key matching (that produced the old 16k false-positive review queue).
#   2. STANDARDIZE each text output column by frequency: cluster near-identical spellings
#      and rewrite the minority to the dominant surface form ("Arijite"×1 → "Arijit"×3).
#   3. Only HEAVY CORRUPTION reaches a human: unrecoverable garble, or a primary key that
#      carries internally contradictory data (same key, conflicting song/artist).
# Everything is in-memory for the single clean run and discarded after — nothing persists.


def build_master(source_files: list[tuple[str | None, bytes]], spec: dict) -> dict:
    """Parse → clean → DEDUP on the primary key. Returns a JSON-friendly result:
    {pk_display, out_cols:[{disp,key,kind}], master:{pk:{canon:val}}, order:[pk,...],
     conflicts:{pk:{canon:[alt,...]}}}.

    The first non-empty value per field wins. When a later row with the SAME key supplies
    a *different* non-empty value for an output column, the alternate is recorded in
    `conflicts` (a same-key data contradiction → candidate for human review)."""
    parsed = []
    for filename, data in source_files:
        headers, body = _parse(filename, data)
        if headers:
            parsed.append((headers, body))
    if not parsed:
        raise ValueError("Could not read any columns from the source files.")

    display: dict[str, str] = {}
    for headers, _ in parsed:
        for h in headers:
            display.setdefault(_norm(h), h)

    pk_norm = _norm(spec.get("primary_key") or "")
    if not pk_norm or pk_norm not in display:
        raise ValueError("The selected primary key is not present in the files.")
    pk_canon = _canon(pk_norm)

    seen, out_cols = set(), []
    for c in (spec.get("columns") or []):
        ck = _canon(c)
        if ck and ck != pk_canon and ck not in seen:
            seen.add(ck)
            if ck in ("release", "golive"):
                kind = "date"
            elif ck == "duration" or _DURATION_COL.search(str(c)):
                kind = "duration"
            else:
                kind = "text"
            out_cols.append({"disp": str(c).strip(), "key": ck, "kind": kind})
    out_keys = {c["key"] for c in out_cols}

    master: dict[str, dict] = {}
    order: list[str] = []
    conflicts: dict[str, dict] = {}
    # A conflict is the SAME source column disagreeing for the same key — compared per
    # source header (not per canonical key), so two distinct fields that happen to alias
    # to one canon ("Album Name" vs "Album cat. No." → album) don't fake a contradiction.
    seen_hdr: dict[str, dict[str, str]] = {}   # header_norm → {pk: first value seen}
    for headers, body in parsed:
        norm_idx = {_norm(h): i for i, h in enumerate(headers)}
        if pk_norm not in norm_idx:
            continue
        pki = norm_idx[pk_norm]
        cols = [(i, _norm(h), _canon(h)) for i, h in enumerate(headers)]
        for row in body:
            key = _clean_text(row[pki] if pki < len(row) else "")
            if not key:
                continue
            if key not in master:
                master[key] = {}
                order.append(key)
            rec = master[key]
            for i, hnorm, ck in cols:
                if ck == pk_canon:
                    continue
                val = _clean_text(row[i] if i < len(row) else "")
                if not val:
                    continue
                if not rec.get(ck):
                    rec[ck] = val
                if ck in out_keys:
                    hs = seen_hdr.setdefault(hnorm, {})
                    prev = hs.get(key)
                    if prev is None:
                        hs[key] = val
                    elif _norm(prev) != _norm(val):
                        alts = conflicts.setdefault(key, {}).setdefault(ck, [])
                        if val not in alts:
                            alts.append(val)

    return {"pk_display": display[pk_norm], "out_cols": out_cols, "master": master,
            "order": order, "conflicts": conflicts}


# ---- Field-value canonicalization (case / diacritics / order / punctuation) ----

def _canon_value(s: str) -> str:
    """Order- and accent-insensitive key for a field value. Multi-value fields ("A & B",
    "B, A") collapse to the same key so spelling votes and conflict checks aren't fooled
    by ordering or diacritics."""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c)).lower()
    s = re.sub(r"\b(feat|ft|featuring|late|shri|sri|smt|dr|mr|mrs)\b", " ", s)
    toks = []
    # split on every common separator incl. dash, so "Shankar-Ehsaan-Loy" == "Shankar,Ehsaan,Loy"
    for p in re.split(r"[&,/\-]| and ", s):
        # drop every non-alphanumeric inside a name so "P V"/"PV"/"P. V." collapse together
        p = re.sub(r"[^a-z0-9]", "", p)
        if p:
            toks.append(p)
    return " ".join(sorted(toks))


# ---- Step 4/5: fuzzy duplicate detection + confidence scoring ----

def _sim(a: str, b: str) -> float:
    a, b = _norm(a), _norm(b)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    return SequenceMatcher(None, a, b).ratio()


def _phonetic(canon: str) -> str:
    """Cheap blocking key for typo clustering — first 4 alphanumerics of the canonical
    form. Keeps the fuzzy spelling pass O(n): only same-block surfaces are compared."""
    return re.sub(r"[^a-z0-9]", "", canon)[:4]


def _standardize_column(master: dict, order: list, ck: str) -> dict:
    """Build a value→canonical map for one text column by frequency.

    Variants that are identical after canonicalization (case/diacritics/order) fold to
    their most frequent surface form. A separate dominance-gated, phonetically-blocked
    pass folds genuine typos ("Arijite"×1 → "Arijit"×3) — only when the target is the
    strictly more frequent spelling, so a rare-but-valid name is never overwritten."""
    counts: Counter = Counter()
    for pk in order:
        v = master[pk].get(ck)
        if v:
            counts[v] += 1
    if len(counts) < 2:
        return {}

    groups: dict[str, Counter] = {}        # canonical key → Counter(surface form)
    for surf, c in counts.items():
        groups.setdefault(_canon_value(surf), Counter())[surf] += c

    rep: dict[str, tuple] = {}             # canonical key → (winning surface, total count)
    value_map: dict[str, str] = {}
    for ckey, sc in groups.items():
        winner = sc.most_common(1)[0][0]
        rep[ckey] = (winner, sum(sc.values()))
        for surf in sc:
            if surf != winner:
                value_map[surf] = winner

    # Typo pass: fold a rare canonical group into a similar, dominant one. Only minority
    # spellings (rare) are folded, and only against the block's most-frequent forms — so
    # it stays ~O(n) (each rare form compared to ≤25 targets) instead of O(distinct²).
    blocks: dict[str, list] = {}
    for ckey, (winner, tot) in rep.items():
        if len(ckey) >= 3:
            blocks.setdefault(_phonetic(ckey), []).append((ckey, winner, tot))
    for group in blocks.values():
        if len(group) < 2:
            continue
        group.sort(key=lambda g: -g[2])             # most frequent first
        targets = [g for g in group if g[2] >= 2][:25]
        if not targets:
            continue
        for ckey, winner, tot in group:
            if tot >= 2:
                continue                             # only fold minority spellings
            for cj, sj, tj in targets:
                if tj > tot and abs(len(ckey) - len(cj)) <= 2 and _sim(ckey, cj) >= 0.90:
                    target = value_map.get(sj, sj)
                    for surf in groups[ckey]:
                        if surf != target:
                            value_map[surf] = target
                    break
    return value_map


def standardize(result: dict, corrections: list | None = None) -> None:
    """In-place: canonicalize spelling across every text output column by frequency vote.
    Generalises to any preset / custom column set (runs per chosen column). Records each
    change in `corrections` for the audit trail."""
    master, order, out_cols = result["master"], result["order"], result["out_cols"]
    for col in out_cols:
        if col["kind"] != "text":
            continue
        ck = col["key"]
        vmap = _standardize_column(master, order, ck)
        if not vmap:
            continue
        for pk in order:
            v = master[pk].get(ck)
            if v and v in vmap:
                master[pk][ck] = vmap[v]
                if corrections is not None and len(corrections) < 5000:
                    corrections.append({"pk": pk, "col": col["disp"], "from": v, "to": vmap[v]})


# ---- Heavy-corruption detection — the only thing escalated to a human ----

# A same-key contradiction is only "heavy corruption" on an *identity* field — the same
# id describing a different song/performer. Drift on metadata (album/date/label/genre)
# is auto-resolved (first value kept) and never shown, since it's mostly format noise.
IDENTITY_KEYS = ("track", "singer", "composer")


def _garbage(v: str) -> bool:
    """A value that is mostly non-alphanumeric (Unicode-aware, so native scripts pass)."""
    if len(v) <= 3:
        return False
    return sum(1 for c in v if c.isalnum()) / len(v) < 0.4


def _real_conflict(cur: str, alts: list, ck: str) -> list:
    """Of the alternate values seen for the same key, return only the ones that are a
    GENUINELY different identity. A subset/superset artist list (extra collaborator) or a
    spelling variant is not heavy corruption — it's auto-resolved, not shown."""
    cur_c = _canon_value(cur)
    out = []
    if ck in ("singer", "composer"):
        cs = set(cur_c.split())
        for a in alts:
            ac = _canon_value(a)
            as_ = set(ac.split())
            if cs and as_ and (cs <= as_ or as_ <= cs):
                continue                       # fuller/looser credit, same people
            if _sim(cur_c, ac) >= 0.85:
                continue                       # spelling variant
            sh, lo = sorted((cur_c, ac), key=len)
            if len(sh) >= 5 and lo.startswith(sh):
                continue                       # first name vs full name (Rupankar / Rupankar Bagchi)
            if ac != cur_c:
                out.append(a)
    else:                                      # track title
        cur_t = _canon_title(cur)
        cur_set = set(cur_t.split())
        for a in alts:
            at = _canon_title(a)
            if _sim(cur_t, at) >= 0.85:
                continue                       # same title (spelling/format)
            aset = set(at.split())
            if len(cur_set) >= 2 and len(aset) >= 2 and (cur_set <= aset or aset <= cur_set):
                continue                       # one title contains the other → same song
            out.append(a)
    return out


def _canon_title(s: str) -> str:
    """Title key with annotations removed — `(From "Movie")`, `[Live]`, a trailing
    `- Singer`/`- Club Mix` version tag — so the same song labelled differently isn't
    read as a different title."""
    s = re.sub(r"[\(\[].*?[\)\]]", " ", str(s))
    s = re.sub(r"\bfrom\b.*$", " ", s, flags=re.I)
    s = re.sub(r"[-–][^-–]*$", " ", s)         # trailing "- <version/singer>"
    return _canon_value(s)


def detect_corruption(result: dict) -> list[dict]:
    """Flag records that genuinely need a human: unrecoverable garble (a replacement
    char ⇒ bytes were lost) or junk values, plus same-key data contradictions recorded
    during the merge. Recoverable issues (mojibake, control chars, ordering, minor
    spelling) are auto-fixed elsewhere and never shown."""
    master, order, out_cols = result["master"], result["order"], result["out_cols"]
    pk_display = result["pk_display"]
    conflicts = result.get("conflicts", {})
    key2disp = {c["key"]: c["disp"] for c in out_cols}
    kind_of = {c["key"]: c["kind"] for c in out_cols}
    scan = [(pk_display, "__pk__")] + [(c["disp"], c["key"]) for c in out_cols]

    def _cmp(val: str, kind: str) -> str:
        # compare on the *cleaned* form so date/duration format and name spacing don't
        # masquerade as a contradiction (12-Jan-26 vs 01/12/2026 → same day).
        return _canon_value(_finish(val, kind))

    review: list[dict] = []
    next_id = 1
    for pk in order:
        rec = master[pk]
        issues: list[dict] = []
        for disp, ck in scan:
            val = pk if ck == "__pk__" else rec.get(ck, "")
            if not val:
                continue
            if "�" in val:
                issues.append({"field": disp, "kind": "garbled", "value": val, "suggestion": ""})
            elif _garbage(val):
                issues.append({"field": disp, "kind": "garbage", "value": val, "suggestion": ""})
        for ck, alts in conflicts.get(pk, {}).items():
            if ck not in IDENTITY_KEYS:
                continue                  # metadata drift → auto-kept, not human-reviewed
            cur = rec.get(ck, "")
            real = _real_conflict(cur, alts, ck)
            if real:
                issues.append({"field": key2disp.get(ck, ck), "kind": "conflict",
                               "value": cur, "suggestion": real[0], "alternates": real})
        if issues:
            review.append({"id": next_id, "pk": pk, "status": "pending",
                           "action": None, "fixes": {}, "issues": issues})
            next_id += 1
    return review


# ---- Apply review decisions, render and store the master ----

def apply_resolutions(staging: dict) -> dict:
    """Apply the operator's per-record decisions. 'fix' writes the supplied corrected
    values; 'delete' drops the record from the output; anything else (dismiss /
    still-pending) keeps the record as-is."""
    master = dict(staging["master"])
    order = list(staging["order"])
    pk_key = _canon(staging["pk_display"])
    deleted: set[str] = set()

    for item in staging.get("review", []):
        if item.get("status") != "resolved":
            continue
        pk = item.get("pk")
        if pk not in master:
            continue
        action = item.get("action")
        if action == "delete":
            deleted.add(pk)
        elif action == "fix":
            for disp, val in (item.get("fixes") or {}).items():
                ck = _canon(disp)
                if ck == pk_key:      # the primary key is the row identity — never rewrite it
                    continue
                master[pk][ck] = _clean_text(val)

    if deleted:
        order = [pk for pk in order if pk not in deleted]

    return {
        "pk_display": staging["pk_display"],
        "out_cols": staging["out_cols"],
        "master": master,
        "order": order,
    }


def _repair(raw: str) -> str:
    """Recover a value: fix mojibake (ftfy) and strip control characters."""
    s = _fix_text(raw)
    s = "".join(c for c in s if ord(c) >= 32 or c == "\t")
    return " ".join(s.split())


def auto_repair(result: dict, corrections: list | None = None) -> None:
    """In-place: silently fix recoverable damage (mojibake, control chars) on the primary
    key and every output column. Values with a lost-byte marker (�) are left for a human."""
    master, order, out_cols = result["master"], result["order"], result["out_cols"]
    for col in out_cols:
        ck, disp = col["key"], col["disp"]
        for pk in order:
            v = master[pk].get(ck)
            if not v or "�" in v:
                continue
            rv = _repair(v)
            if rv and rv != v:
                master[pk][ck] = rv
                if corrections is not None and len(corrections) < 5000:
                    corrections.append({"pk": pk, "col": disp, "from": v, "to": rv})


def clean_and_detect(source_files: list[tuple[str | None, bytes]], spec: dict) -> tuple[dict, list[dict]]:
    """Dedup on the primary key → auto-repair → standardize spelling → flag heavy
    corruption. CPU-bound; call via asyncio.to_thread from async routes. Returns
    (result, review) where review holds only the records a human must look at."""
    result = build_master(source_files, spec)
    corrections: list[dict] = []
    auto_repair(result, corrections)
    standardize(result, corrections)
    review = detect_corruption(result)
    result["corrections"] = corrections
    result["corrections_total"] = len(corrections)
    return result, review


def display_record(out_cols: list[dict], rec: dict) -> dict:
    """Cleaned, display-ready values for one record, keyed by output column name."""
    return {c["disp"]: _finish(rec.get(c["key"], ""), c["kind"]) for c in out_cols}


def split_corrupted(staging: dict) -> tuple[list, list, dict]:
    """Partition the staged master into clean vs flagged rows for the 'skip & download'
    export. Returns (good_pks, corrupted_pks, issue_by_pk)."""
    review = staging.get("review", [])
    corrupted = {it["pk"] for it in review}
    issue_by_pk = {
        it["pk"]: "; ".join(f'{i["field"]} ({i["kind"]})' for i in it.get("issues", []))
        for it in review
    }
    good = [pk for pk in staging["order"] if pk not in corrupted]
    corr = [pk for pk in staging["order"] if pk in corrupted]
    return good, corr, issue_by_pk


def render_csv(result: dict) -> bytes:
    out = io.StringIO(newline="")
    w = csv.writer(out)
    cols = result["out_cols"]
    w.writerow([result["pk_display"]] + [c["disp"] for c in cols])
    for pk in result["order"]:
        rec = result["master"].get(pk, {})
        w.writerow([pk] + [_finish(rec.get(c["key"], ""), c["kind"]) for c in cols])
    return out.getvalue().encode("utf-8")


def _finish(value: str, kind: str) -> str:
    if kind == "date":
        return _norm_date(value)
    if kind == "duration":
        return _norm_duration(value)
    return value


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def render_xlsx(result: dict, pks: list | None = None, issue_by_pk: dict | None = None) -> bytes:
    """Render the master (or a subset of pks) to an .xlsx workbook. When `issue_by_pk` is
    given, an extra 'Issue' column describes why each row was flagged (corrupted export)."""
    from openpyxl import Workbook

    cols = result["out_cols"]
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    header = [result["pk_display"]] + [c["disp"] for c in cols]
    if issue_by_pk is not None:
        header.append("Issue")
    ws.append(header)
    for pk in (pks if pks is not None else result["order"]):
        rec = result["master"].get(pk, {})
        row = [pk] + [_finish(rec.get(c["key"], ""), c["kind"]) for c in cols]
        if issue_by_pk is not None:
            row.append(issue_by_pk.get(pk, ""))
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def store_output(
    branch: Branch, data: bytes, storage: StorageBackend | None,
    *, kind: str = "cleaned", suffix: str = "cleaned.csv", mime: str = "text/csv",
) -> File:
    """Persist an output blob as a File row (Drive object or DB bytes per backend)."""
    out_name = f"{branch.name.replace(' ', '_')}_{suffix}"
    if settings.STORAGE_BACKEND == "drive":
        if storage is None:
            raise ValueError("Drive storage is not configured.")
        key = await asyncio.to_thread(storage.put, data, out_name, mime)
        return File(
            branch_id=branch.id, kind=kind, storage_key=key, content=None,
            content_bytes=None, original_filename=out_name, mime_type=mime,
            size_bytes=len(data), status="available",
        )
    return File(
        branch_id=branch.id, kind=kind, storage_key="", content=None,
        content_bytes=data, original_filename=out_name, mime_type=mime,
        size_bytes=len(data), status="available",
    )


async def store_cleaned(branch: Branch, cleaned_bytes: bytes, storage: StorageBackend | None) -> File:
    return await store_output(branch, cleaned_bytes, storage, kind="cleaned",
                              suffix="cleaned.csv", mime="text/csv")
