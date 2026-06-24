"""Standardize a messy input file into the canonical 30-column master format.

This is the engine behind the "Standardize" tab. Unlike the cleaning pipeline it
does NOT flag, score or queue anything for human review — it just:

  1. Allocates every input column to the right master column (auto-mapped), with
     several input columns feeding one master column merged with " | " (e.g.
     "Singer 1" / "Singer 2" -> Singer, or "Composer" / "Music Director" ->
     Composer).
  2. Explodes packed "role: name | role: name" credit columns (audioCredits,
     musicianCredits, ...) into per-role virtual columns first, so the singer /
     composer / lyricist / label buried inside one cell land in their own master
     columns instead of being jumbled together.
  3. Applies the same value standardization the cleaner uses (Title-cased names,
     ISO dates, upper-cased ISRC, digit-only UPC, mm:ss durations, ...), but only
     keeps the standardized value — every cell is accepted as-is, nothing is
     marked an error.

The output is always all 30 master columns, in canonical order; columns that
nothing maps to come out blank, ready to be filled later by the cleaning flow.
"""

import csv
import io
import re

from openpyxl import load_workbook

from ..models import MASTER_COLUMN_TO_ATTR
from .cleaning import NAME_SEP, clean_value
from .matching import SAMPLE_SIZE, _header_score, suggest_mapping

# Standardize runs with NO human review, so it errs on the side of precision:
# a column is only allocated when the evidence is strong. A weak fuzzy guess is
# left unmapped (the master column comes out blank, to be filled later in the
# cleaning flow) rather than risking the data landing in the wrong column.
#   - A fuzzy header match must clear this to be accepted as a primary source.
#   - An extra (secondary) source must be a genuine NAME variant of the column
#     (e.g. "Singer 2" alongside "Singer"), judged on header similarity alone —
#     this is what stops a same-shaped-but-unrelated column (a stray numeric
#     field looking like a percentage) from being merged in on content evidence.
_STRICT = 0.85

# The canonical output schema, in order.
MASTER_COLUMNS: list[str] = list(MASTER_COLUMN_TO_ATTR)

# --------------------------------------------------------------------------
# Packed "credit" columns: "composer: A | singer: B | lyricist: C"
# --------------------------------------------------------------------------
# One "role: value" segment. The role is a short label (letters, spaces and a
# few joiners); the value is everything after the first colon.
_SEGMENT = re.compile(r"^\s*([A-Za-z][A-Za-z ./&-]{0,30}?)\s*:\s*(.+?)\s*$")
# Segments inside a packed cell are joined by pipes / semicolons / newlines.
_PACK_SPLIT = re.compile(r"\s*[|;\n]+\s*")
# A column is treated as "packed" when at least this fraction of its non-empty
# sample cells parse into 2+ role:value segments.
_PACK_MIN = 0.6


def _segments(cell: str) -> list[tuple[str, str]]:
    """The (role, value) pairs in one packed cell (role lower-cased)."""
    out: list[tuple[str, str]] = []
    for part in _PACK_SPLIT.split(str(cell or "")):
        m = _SEGMENT.match(part)
        if not m:
            continue
        role = re.sub(r"\s+", " ", m.group(1).strip().lower())
        value = m.group(2).strip()
        if role and value:
            out.append((role, value))
    return out


def _looks_packed(values: list) -> bool:
    """True when a column's cells are mostly "role: value | role: value" blobs."""
    hits = total = 0
    for v in values:
        s = str(v or "").strip()
        if not s:
            continue
        total += 1
        if len(_segments(s)) >= 2:
            hits += 1
        if total >= SAMPLE_SIZE:
            break
    return total > 0 and hits / total >= _PACK_MIN


def _explode_packed(
    headers: list[str], rows: list[list]
) -> tuple[list[str], list[list], list[str]]:
    """Turn packed credit columns into per-role virtual columns.

    Returns augmented (headers, rows, new_role_headers). Each distinct role found
    across the packed columns becomes a new synthetic column named after the role
    (e.g. "singer", "composer", "lyricist", "label"); the auto-mapper then routes
    those to the matching master columns by name. Roles with no master match
    (producer, mixed by, ...) simply stay unmapped and drop out of the output.

    The original packed columns are REMOVED from the augmented table: their raw
    "role: name | role: name" blob has been fully decomposed into the role
    columns, so leaving it in would only let the matcher dump the whole blob into
    some unrelated catch-all (path/text) column.
    """
    cols: list[list] = [[] for _ in headers]
    for r in rows:
        for i in range(len(headers)):
            cols[i].append(r[i] if i < len(r) else "")
    packed = {i for i in range(len(headers)) if _looks_packed(cols[i])}
    if not packed:
        return headers, rows, []

    kept = [i for i in range(len(headers)) if i not in packed]
    existing = {re.sub(r"\s+", " ", h.strip().lower()) for h in headers}
    roles: list[str] = []
    for i in packed:
        for cell in cols[i]:
            for role, _ in _segments(cell):
                if role not in roles and role not in existing:
                    roles.append(role)

    role_pos = {role: idx for idx, role in enumerate(roles)}
    new_rows: list[list] = []
    for r in rows:
        bucket: list[list[str]] = [[] for _ in roles]
        for i in packed:
            cell = r[i] if i < len(r) else ""
            for role, value in _segments(cell):
                pos = role_pos.get(role)
                if pos is not None and value not in bucket[pos]:
                    bucket[pos].append(value)
        new_rows.append(
            [r[i] if i < len(r) else "" for i in kept] + [NAME_SEP.join(b) for b in bucket]
        )
    return [headers[i] for i in kept] + roles, new_rows, roles


def _tighten(mapping: list[dict]) -> list[dict]:
    """Drop low-confidence allocations so nothing lands in the wrong column.

    A weak fuzzy primary is unmapped entirely; an extra source is kept only when
    its header is genuinely similar to the master column (a real numbered/variant
    column), never on content evidence alone. Mutates and returns `mapping`.
    """
    for m in mapping:
        if not m.get("input_header"):
            continue
        col = m["master_column"]
        if m.get("method") == "fuzzy" and m.get("confidence", 0.0) < _STRICT:
            m["input_header"] = None
            m["extra_headers"] = []
            m["method"] = "unmatched"
            m["confidence"] = 0.0
            continue
        m["extra_headers"] = [
            h for h in (m.get("extra_headers") or []) if _header_score(h, col)[0] >= _STRICT
        ]
    return mapping


# --------------------------------------------------------------------------
# Per-value pre-normalization (before the shared cleaner runs)
# --------------------------------------------------------------------------
_SECONDS = re.compile(r"^\d+(\.\d+)?$")
# Invisible zero-width / word-joiner / BOM characters that ride along in messy
# exports (e.g. a leading U+2060 on a title) and survive NFKC normalization.
_INVISIBLE = re.compile("[\u200b\u200c\u200d\u2060\ufeff]")


def _pre_normalize(master_column: str, raw):
    """Light, type-aware fix-ups the generic cleaner can't infer on its own.

      - Strip invisible zero-width / BOM characters from every value.
      - A bare numeric Audio Duration is read as a count of seconds
        ("352.11" -> "5:52") so the cleaner standardizes it to mm:ss instead of
        rejecting it.
    """
    if raw is None:
        return raw
    s = _INVISIBLE.sub("", str(raw))
    if master_column == "Audio Duration (mm:sec)" and _SECONDS.match(s.strip()):
        total = int(round(float(s.strip())))
        minutes, seconds = divmod(total, 60)
        return f"{minutes}:{seconds:02d}"
    return s


# --------------------------------------------------------------------------
# Standardization
# --------------------------------------------------------------------------
def standardize(headers: list[str], rows: list[list]) -> dict:
    """Map + standardize an input table into the master format.

    Returns {columns, rows, mapping} where `columns` is the full master schema,
    `rows` are dicts keyed by master column, and `mapping` describes how each
    master column was filled (its source input columns).
    """
    aug_headers, aug_rows, _roles = _explode_packed(headers, rows)
    sample = aug_rows[:SAMPLE_SIZE]
    mapping = _tighten(suggest_mapping(aug_headers, MASTER_COLUMNS, sample)["mappings"])
    active = [m for m in mapping if m.get("input_header")]

    hidx = {h: i for i, h in enumerate(aug_headers)}
    out_rows: list[dict] = []
    for n, r in enumerate(aug_rows, start=1):
        values = {c: "" for c in MASTER_COLUMNS}
        for m in active:
            col = m["master_column"]
            sources = [m["input_header"], *(m.get("extra_headers") or [])]
            raws = []
            for h in sources:
                i = hidx.get(h)
                cell = r[i] if (i is not None and i < len(r)) else None
                raws.append(_pre_normalize(col, cell))
            values[col] = clean_value(col, raws).value
        # Record # is positional: keep an input value, else number rows 1..N.
        if not values["Record #"]:
            values["Record #"] = str(n)
        out_rows.append(values)

    return {"columns": MASTER_COLUMNS, "rows": out_rows, "mapping": mapping}


def mapping_summary(mapping: list[dict]) -> list[dict]:
    """A compact per-master-column view of the resolved mapping for the UI:
    {master_column, sources, matched}. Only the mapped columns carry sources."""
    out = []
    for m in mapping:
        primary = m.get("input_header")
        sources = [primary, *(m.get("extra_headers") or [])] if primary else []
        out.append(
            {
                "master_column": m["master_column"],
                "sources": [s for s in sources if s],
                "matched": bool(primary),
            }
        )
    return out


# --------------------------------------------------------------------------
# Input loading: .xlsx / .xlsm / .csv -> (headers, rows of strings)
# --------------------------------------------------------------------------
class StandardizeError(Exception):
    def __init__(self, message: str):
        self.message = message
        super().__init__(message)


def _cell_to_str(v) -> str:
    import datetime as _dt

    if v is None:
        return ""
    if isinstance(v, _dt.datetime):
        return v.isoformat(sep=" ")
    if isinstance(v, (_dt.date, _dt.time)):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))  # keep barcodes from rendering as 8.9e+12
    return str(v)


def _load_xlsx(data: bytes) -> tuple[list[str], list[list]]:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises various errors on bad files
        raise StandardizeError(
            f"The file could not be opened ({type(exc).__name__})."
        )
    ws = wb.active
    if ws is None:
        raise StandardizeError("The workbook has no sheets.")

    header: list[str] | None = None
    rows: list[list] = []
    for raw in ws.iter_rows(values_only=True):
        if not any(v not in (None, "") for v in raw):
            continue
        if header is None:
            header = [str(v).strip() if v is not None else "" for v in raw]
            while header and header[-1] == "":
                header.pop()
            continue
        n = len(header)
        cells = list(raw[:n]) + [None] * (n - len(raw[:n]))
        rows.append([_cell_to_str(v) for v in cells])
    wb.close()
    if not header:
        raise StandardizeError("No column headers were found.")
    return header, rows


def _load_csv(data: bytes) -> tuple[list[str], list[list]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    header: list[str] | None = None
    rows: list[list] = []
    for raw in reader:
        if not any((c or "").strip() for c in raw):
            continue
        if header is None:
            header = [c.strip() for c in raw]
            while header and header[-1] == "":
                header.pop()
            continue
        n = len(header)
        cells = list(raw[:n]) + [""] * (n - len(raw[:n]))
        rows.append([c if c is not None else "" for c in cells])
    if not header:
        raise StandardizeError("No column headers were found.")
    return header, rows


def load_table(filename: str, data: bytes) -> tuple[list[str], list[list]]:
    """Parse an uploaded .csv / .xlsx / .xlsm into (headers, string rows)."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        headers, rows = _load_csv(data)
    elif name.endswith((".xlsx", ".xlsm")):
        headers, rows = _load_xlsx(data)
    else:
        raise StandardizeError("Only .csv, .xlsx and .xlsm files are supported.")
    if not rows:
        raise StandardizeError("The file has headers but no data rows.")
    return headers, rows
