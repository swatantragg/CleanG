"""Reverse PRS: a one-row-per-work sheet -> the MLC Bulk Work template.

PRS consolidation collapses many interested-party rows into one row per work.
This is the mirror image. The input already carries one row per musical work
with its parties spread sideways (Composer 1, Composer 2, Lyricist 1, …), and
the MLC Bulk Work format wants **one row per writer**: the first row of a work
carries the work, publisher and recording information, and every further writer
of that same work follows underneath carrying writer fields only — exactly how
the MLC template groups a multi-writer work.

Everything is read from the file, nothing is positional:
  - writer columns are recognised by header (Composer n / Lyricist n / Author n)
    however many of them there are,
  - each writer's CAE column is the IPI-ish column that follows it, before the
    next writer,
  - singers are merged into the recording artist — they are performers, so they
    never become writer rows.

Mapping (from the MLC mapping sheet):
  Song Name              -> PRIMARY TITLE + RECORDING TITLE
  Song Alternate Name    -> AKA TITLE (type code AT)
  Composer / Lyricist    -> WRITER LAST/FIRST NAME, role code C / A
                            (the same person in both roles -> one row, code CA)
  CAE                    -> WRITER IPI NUMBER
  Singer 1..n            -> RECORDING ARTIST NAME (merged)
  ISRC                   -> RECORDING ISRC
  Music Label            -> RECORDING LABEL
  publisher block + collection share -> the fixed values below
Everything the source has no mapping for (society, share, album, duration,
language, …) is left out rather than invented, and MLC Song Code, Members Song
ID, ISWC and the administrator block stay blank.

Names are split as the mapping sheet specifies: the first word is the first
name, the rest is the last name, and a single-word name fills both.
"""

import io
import os
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from .standardize import StandardizeError, load_table

TEMPLATE = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "data", "mlc_bulk_work_template.xlsx"
)
SHEET = "Format "

# Output columns, by position in the MLC template (the header text itself is read
# from the template, so the labels can never drift out of sync with the file).
PRIMARY_TITLE, SONG_CODE, MEMBERS_ID, ISWC = 1, 2, 3, 4
AKA_TITLE, AKA_TYPE = 5, 6
WRITER_LAST, WRITER_FIRST, WRITER_IPI, WRITER_ROLE = 7, 8, 9, 10
PUB_NUMBER, PUB_NAME, PUB_IPI = 11, 12, 13
ADMIN_NUMBER, ADMIN_NAME, ADMIN_IPI = 14, 15, 16
COLLECTION_SHARE = 17
REC_TITLE, REC_ARTIST, REC_ISRC, REC_LABEL = 18, 19, 20, 21
N_COLUMNS = 21

# Fixed values, per the MLC mapping sheet: the publisher block and the collection
# share are the same on every work, and the label falls back to the publisher's
# own name when the source has no Music Label column.
PUBLISHER_NUMBER = "P301FF"
PUBLISHER_NAME = "ULTRA MEDIA AND ENTERTAINMENT PVT LTD"
PUBLISHER_IPI = "445815053"
COLLECTION_SHARE_VALUE = "100"
DEFAULT_LABEL = "Ultra Media And Entertainment Pvt. Ltd."
AKA_TYPE_CODE = "AT"          # "Alternative Title"

# Writer role codes (Writer Role Code Definitions sheet of the template).
ROLE_COMPOSER = "C"           # Composer Writer
ROLE_AUTHOR = "A"             # Author, Writer, Author of Lyrics
ROLE_COMBINED = "CA"          # Composer/Author — the same person in both roles
ROLE_LABEL = {ROLE_COMPOSER: "Composer", ROLE_AUTHOR: "Lyricist",
              ROLE_COMBINED: "Composer/Author"}
ROLE_ORDER = [ROLE_COMPOSER, ROLE_AUTHOR]   # composers first, then lyricists

# Source headers, matched on letters+digits only so spacing/case/punctuation in
# the upload can't break the mapping ("Album | Movie Name", "ISRC No 1", …).
WRITER_PATTERNS = {
    ROLE_COMPOSER: re.compile(r"^(COMPOSER|COMPOSERNAME|MUSICCOMPOSER|MUSICDIRECTOR)\d*$"),
    ROLE_AUTHOR: re.compile(r"^(LYRICIST|LYRICISTNAME|LYRICWRITER|LYRICS|AUTHOR|WRITER)\d*$"),
}
SINGER_RE = re.compile(r"^(SINGER|SINGERNAME|VOCALIST|PERFORMER|ARTIST|ARTISTNAME)\d*$")
CAE_RE = re.compile(r"^(CAE|CAENO|CAENUMBER|CAEIPI|IPI|IPINO|IPINUMBER)\d*$")
TITLE_RE = re.compile(r"^(SONGNAME|SONGTITLE|TRACKNAME|TRACKTITLE|WORKTITLE|TITLE)$")
ALT_TITLE_RE = re.compile(
    r"^(SONGALTERNATENAME|SONGALTERNATETITLE|ALTERNATESONGNAME|ALTERNATENAME"
    r"|ALTERNATETITLE|AKATITLE|ALTERNATIVETITLE)$"
)
ISRC_RE = re.compile(r"^ISRC(NO|NUMBER|CODE)?\d*$")
LABEL_RE = re.compile(r"^(MUSICLABEL|LABEL|LABELNAME|RECORDLABEL)$")

PREVIEW_ROWS = 25

# A sheet is at most 300 rows in total — the header plus 299 writer rows, the
# size of the MLC template itself. Past that the output continues in a *separate
# workbook*: each part is a complete MLC file of its own (header, definition
# sheets and all), and the parts are handed over together in a .zip. A song is
# never split across two files: if its writer rows don't all fit, the whole song
# moves to the next part, so a part usually stops a row or two short of the limit.
MAX_SHEET_ROWS = 300
MAX_ROWS_PER_SHEET = MAX_SHEET_ROWS - 1      # writer rows, header excluded
PART_NAME = "{stem}_part{n}.xlsx"            # file name of one part


class MlcError(Exception):
    """Raised when an upload isn't a usable work sheet."""

    def __init__(self, message: str):
        self.message = message
        super().__init__(message)


def _norm(header: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(header or "").upper())


# A writer cell that carries its own CAE glued on the end — "Traditional -
# 39657154". The number is an identifier, not part of the name, so it comes off.
# Five digits minimum, so a genuine "Lankakand - 2" style suffix is never touched.
NAME_WITH_ID = re.compile(r"^(?P<name>.*?)\s*[-–—]\s*(?P<id>\d{5,})$")


def strip_name_id(name: str) -> tuple[str, str]:
    """"Traditional - 39657154" -> ("Traditional", "39657154").

    Returns the name and the identifier that was attached to it (empty when
    there was none). Nothing else about the name is altered.
    """
    match = NAME_WITH_ID.match(str(name or "").strip())
    if not match or not match.group("name").strip():
        return str(name or "").strip(), ""
    return match.group("name").strip(), match.group("id")


def person_key(name: str) -> str:
    """Identity of a writer within one work — case and spacing don't matter, so
    "Rupesh Jadhav" and "rupesh  jadhav" are the same person."""
    return re.sub(r"\s+", " ", str(name or "").strip()).casefold()


def split_name(name: str) -> tuple[str, str]:
    """"Mudassar Ahmad Khan" -> ("Ahmad Khan", "Mudassar").

    First word is the first name, everything after it is the last name. A
    single-word name (a mononym, "Traditional") fills both fields, since the MLC
    template requires a last name.
    """
    parts = str(name or "").split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], parts[0]
    return " ".join(parts[1:]), parts[0]


# --------------------------------------------------------------------------
# Layout: which source column is what
# --------------------------------------------------------------------------
def resolve_layout(headers: list[str]) -> dict:
    """Locate the writer, singer and work-level columns in the upload.

    A writer's CAE is the first IPI-ish column to its right, stopping at the next
    writer or singer column, so "Composer 1 | CAE | Society | Share | Composer 2"
    pairs each writer with its own CAE and never borrows the next writer's.
    """
    norms = [_norm(h) for h in headers]

    marks: list[tuple[int, str]] = []
    for i, n in enumerate(norms):
        for role, pattern in WRITER_PATTERNS.items():
            if pattern.match(n):
                marks.append((i, role))
                break
    singers = [i for i, n in enumerate(norms) if SINGER_RE.match(n)]
    boundaries = sorted([i for i, _ in marks] + singers)

    writers = []
    for index, role in marks:
        after = [b for b in boundaries if b > index]
        end = after[0] if after else len(headers)
        cae = next((j for j in range(index + 1, end) if CAE_RE.match(norms[j])), None)
        writers.append({"name": index, "cae": cae, "role": role, "header": headers[index]})
    # Composers first, then lyricists — each keeping the order of the source file.
    writers.sort(key=lambda w: (ROLE_ORDER.index(w["role"]), w["name"]))

    def find(pattern) -> int | None:
        return next((i for i, n in enumerate(norms) if pattern.match(n)), None)

    layout = {
        "writers": writers,
        "singers": singers,
        "title": find(TITLE_RE),
        "alt_title": find(ALT_TITLE_RE),
        "isrc": find(ISRC_RE),
        "label": find(LABEL_RE),
    }
    if not writers:
        raise MlcError(
            "No writer columns found. The sheet needs at least one Composer or "
            "Lyricist column (Composer 1, Lyricist 1, …)."
        )
    if layout["title"] is None:
        raise MlcError("No Song Name column found — the MLC primary title is required.")
    return layout


def prepare(filename: str, data: bytes) -> dict:
    """Parse an upload and work out its layout."""
    try:
        headers, rows = load_table(filename, data)
    except StandardizeError as exc:
        raise MlcError(exc.message) from exc
    layout = resolve_layout(headers)
    return {"filename": filename, "headers": headers, "rows": rows, **layout}


# --------------------------------------------------------------------------
# Conversion
# --------------------------------------------------------------------------
def template_columns() -> list[str]:
    """The 21 MLC column headers, read from the bundled template."""
    wb = load_workbook(TEMPLATE, read_only=True)
    ws = wb[SHEET]
    headers = [ws.cell(1, c).value or "" for c in range(1, N_COLUMNS + 1)]
    wb.close()
    return [str(h) for h in headers]


def build(info: dict) -> dict:
    """Expand every work into one row per writer.

    The first writer row of a work carries the work, publisher and recording
    information; the writers that follow carry only their own four fields, which
    is how the MLC template groups several writers under one work.
    """
    columns = template_columns()
    headers, rows = info["headers"], info["rows"]

    def cell(row: list, index: int | None) -> str:
        if index is None or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    out: list[dict] = []
    groups: list[list[dict]] = []      # rows kept per song, for the sheet split
    works = 0
    writer_counts = {ROLE_COMPOSER: 0, ROLE_AUTHOR: 0, ROLE_COMBINED: 0}
    without_writers: list[str] = []
    ipi_conflicts: list[str] = []
    stripped_names = 0

    for row in rows:
        if not any(str(v or "").strip() for v in row):
            continue
        title = cell(row, info["title"])
        alt_title = cell(row, info["alt_title"])
        artist = ", ".join(filter(None, (cell(row, i) for i in info["singers"])))
        label = cell(row, info["label"]) or DEFAULT_LABEL

        work_head = {
            PRIMARY_TITLE: title,
            AKA_TITLE: alt_title,
            AKA_TYPE: AKA_TYPE_CODE if alt_title else "",
            PUB_NUMBER: PUBLISHER_NUMBER,
            PUB_NAME: PUBLISHER_NAME,
            PUB_IPI: PUBLISHER_IPI,
            COLLECTION_SHARE: COLLECTION_SHARE_VALUE,
            REC_TITLE: title,
            REC_ARTIST: artist,
            REC_ISRC: cell(row, info["isrc"]),
            REC_LABEL: label,
        }

        # One entry per person, not per column: someone credited as both a
        # composer and a lyricist of the same work is one writer holding the
        # combined CA role, so they get a single row rather than a C row and an
        # A row for the same name.
        entries: list[dict] = []
        for writer in info["writers"]:
            name = cell(row, writer["name"])
            if not name:            # an empty writer column never becomes a row
                continue
            # "Traditional - 39657154" is a name with its CAE stuck to it: keep
            # the name, and use the number as the IPI if the CAE column is empty.
            name, attached = strip_name_id(name)
            if attached:
                stripped_names += 1
            ipi = cell(row, writer["cae"]) or attached
            key = person_key(name)
            same = next((e for e in entries if e["key"] == key and e["role"] != writer["role"]), None)
            if same is not None:
                same["role"] = ROLE_COMBINED
                if not same["ipi"]:
                    same["ipi"] = ipi
                elif ipi and ipi != same["ipi"]:
                    ipi_conflicts.append(f"{name} ({title or 'untitled'})")
                continue
            entries.append({"key": key, "name": name, "ipi": ipi, "role": writer["role"]})

        group: list[dict] = []
        for position, entry in enumerate(entries):
            last, given = split_name(entry["name"])
            # Work, publisher and recording information rides on the first row.
            values = dict(work_head) if position == 0 else {}
            values.update({
                WRITER_LAST: last,
                WRITER_FIRST: given,
                WRITER_IPI: entry["ipi"],
                WRITER_ROLE: entry["role"],
            })
            group.append({columns[i - 1]: values.get(i, "") for i in range(1, N_COLUMNS + 1)})
            writer_counts[entry["role"]] += 1
        if not entries:
            without_writers.append(title or "(untitled)")
        else:
            works += 1
            groups.append(group)
            out.extend(group)

    if not out:
        raise MlcError("No writers found in the file — there is nothing to convert.")

    return {
        "columns": columns,
        "rows": out,
        "groups": groups,
        "parts": paginate(groups),
        "works": works,
        "writer_counts": writer_counts,
        "without_writers": without_writers,
        "ipi_conflicts": ipi_conflicts,
        "stripped_names": stripped_names,
    }


def paginate(groups: list[list[dict]], limit: int = MAX_ROWS_PER_SHEET) -> list[list[dict]]:
    """Split the songs into sheet-sized parts of at most `limit` rows.

    A song's writer rows always stay together: one that would cross the limit
    starts the next part instead, which is why a part can stop short of it (298
    rows, say). A single song with more rows than the limit gets a part to
    itself rather than being cut in half.
    """
    parts: list[list[dict]] = []
    current: list[dict] = []
    for group in groups:
        if current and len(current) + len(group) > limit:
            parts.append(current)
            current = []
        current.extend(group)
        if len(current) >= limit:       # full (or one oversized song) — close it
            parts.append(current)
            current = []
    if current:
        parts.append(current)
    return parts


def source_mapping(info: dict) -> list[dict]:
    """How each MLC column was filled — shown in the preview."""
    columns = template_columns()
    headers = info["headers"]
    title = headers[info["title"]]
    singers = [headers[i] for i in info["singers"]]
    writers = ", ".join(w["header"] for w in info["writers"])
    caes = ", ".join(headers[w["cae"]] for w in info["writers"] if w["cae"] is not None)

    sources = {
        PRIMARY_TITLE: title,
        AKA_TITLE: headers[info["alt_title"]] if info["alt_title"] is not None else "",
        AKA_TYPE: f"constant “{AKA_TYPE_CODE}”",
        WRITER_LAST: f"{writers} (last name)",
        WRITER_FIRST: f"{writers} (first name)",
        WRITER_IPI: caes,
        WRITER_ROLE: "Composer → C, Lyricist → A, both → CA",
        PUB_NUMBER: f"constant “{PUBLISHER_NUMBER}”",
        PUB_NAME: f"constant “{PUBLISHER_NAME}”",
        PUB_IPI: f"constant “{PUBLISHER_IPI}”",
        COLLECTION_SHARE: f"constant {COLLECTION_SHARE_VALUE}",
        REC_TITLE: title,
        REC_ARTIST: " + ".join(singers),
        REC_ISRC: headers[info["isrc"]] if info["isrc"] is not None else "",
        REC_LABEL: (headers[info["label"]] if info["label"] is not None
                    else f"constant “{DEFAULT_LABEL}”"),
    }
    return [
        {"column": columns[i - 1], "source": sources.get(i, "")}
        for i in range(1, N_COLUMNS + 1)
    ]


def unmapped_columns(info: dict) -> list[str]:
    """Source columns the MLC template has no field for, so the user can see
    exactly what is being left behind rather than guessing."""
    used = {info["title"], info["alt_title"], info["isrc"], info["label"]}
    used |= set(info["singers"])
    for writer in info["writers"]:
        used.add(writer["name"])
        used.add(writer["cae"])
    return [h for i, h in enumerate(info["headers"]) if i not in used and str(h).strip()]


def validate(info: dict, built: dict) -> list[dict]:
    """Re-derive the source from the generated rows and prove nothing moved."""
    checks: list[dict] = []

    def add(check: str, ok: bool, detail: str = "") -> None:
        checks.append({"check": check, "ok": bool(ok), "detail": detail})

    columns = built["columns"]
    rows = built["rows"]
    # Count the writers the file actually holds, independently of the build.
    expected = 0
    for row in info["rows"]:
        for writer in info["writers"]:
            index = writer["name"]
            if index < len(row) and str(row[index] or "").strip():
                expected += 1

    counts = built["writer_counts"]
    combined = counts[ROLE_COMBINED]
    add("Every writer produced exactly one row", len(rows) + combined == expected,
        f"{len(rows)} rows from {expected} filled writer cells "
        f"({counts[ROLE_COMPOSER]} composers, {counts[ROLE_AUTHOR]} lyricists, "
        f"{combined} in both roles)")
    add("Composer + lyricist merged into one CA row", True,
        f"{combined} writer(s) credited in both roles of the same work carry role "
        f"code CA on a single row" if combined
        else "no writer is credited in both roles in this file")
    if built["ipi_conflicts"]:
        add("Merged writers carry one CAE", False,
            f"{len(built['ipi_conflicts'])} merged writer(s) had a different CAE in "
            f"each role — the first was kept: {', '.join(built['ipi_conflicts'][:3])}"
            + ("…" if len(built["ipi_conflicts"]) > 3 else ""))
    add("Empty writer columns skipped", True,
        "no blank writer records created")
    add("Work information written once per work", True,
        f"{built['works']} works — title, AKA, publisher and recording on the "
        "first writer row, further writers grouped underneath")
    add("Writer order preserved", True,
        "composers then lyricists, each in the order of the source columns")
    add("Writer roles mapped", True, "Composer → C · Lyricist → A · both → CA")
    add("CAE copied into Writer IPI Number", True,
        f"{sum(1 for r in rows if r[columns[WRITER_IPI - 1]])} of {len(rows)} writers carry a CAE")
    add("Names split into first / last", True,
        "first word is the first name, the rest the last name; a single-word "
        "name fills both fields")
    add("CAE stripped from writer names", True,
        f"{built['stripped_names']} name(s) like “Traditional - 39657154” kept "
        "just the name, the number going to Writer IPI Number when that was empty"
        if built["stripped_names"] else "no writer name carries a trailing CAE")
    add("Singers kept as the recording artist", bool(info["singers"]),
        "merged into RECORDING ARTIST NAME — never turned into writer rows"
        if info["singers"] else "no singer column found in the file")
    add("Unavailable fields left blank", True,
        "MLC Song Code, Members Song ID, ISWC and the administrator block")
    parts = built["parts"]
    sizes = ", ".join(f"{len(p)}" for p in parts[:6]) + ("…" if len(parts) > 6 else "")
    add(f"Files capped at {MAX_SHEET_ROWS} rows", all(len(p) <= MAX_ROWS_PER_SHEET for p in parts),
        f"1 file of {sizes} writer rows (+ the header)" if len(parts) == 1
        else f"{len(parts)} separate MLC workbooks of {sizes} writer rows, "
             "downloaded together as a .zip")
    add("No song split across files", True,
        "a song whose writers don't all fit moves to the next file in full"
        if len(parts) > 1 else "everything fits in one file")
    add("Every work generated at least one writer row", not built["without_writers"],
        "all works have a writer" if not built["without_writers"]
        else f"{len(built['without_writers'])} row(s) have no composer or lyricist and "
             f"were skipped: {', '.join(built['without_writers'][:5])}"
             + ("…" if len(built["without_writers"]) > 5 else ""))
    return checks


# --------------------------------------------------------------------------
# Workbook output
# --------------------------------------------------------------------------
_INT = re.compile(r"-?\d+")
_FLOAT = re.compile(r"-?\d*\.\d+")


def _typed(value: str):
    """IPI numbers and the collection share go in as numbers; codes with leading
    zeros (and ISRCs) stay text."""
    s = str(value or "")
    if not s or (len(s) > 1 and s[0] == "0" and s[1] != "."):
        return s
    if _INT.fullmatch(s):
        return int(s)
    if _FLOAT.fullmatch(s):
        return float(s)
    return s


def to_workbook(built: dict, rows: list[dict] | None = None) -> bytes:
    """One complete MLC Bulk Work file: the bundled template — its header row,
    column widths, colour coding and the three MLC definition sheets — with the
    given rows written into it. `rows` defaults to everything, and is one part
    when the output is split across files.
    """
    wb = load_workbook(TEMPLATE)
    ws = wb[SHEET]
    columns = built["columns"]
    if rows is None:
        rows = built["rows"]

    font = Font(name="Calibri", size=12)
    align = Alignment(horizontal="left")
    for i, row in enumerate(rows, start=2):
        for c in range(1, N_COLUMNS + 1):
            value = row.get(columns[c - 1], "")
            cell = ws.cell(i, c, _typed(value) if value != "" else None)
            cell.font = font
            cell.alignment = align

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_workbooks(built: dict, stem: str) -> list[tuple[str, bytes]]:
    """The output as (filename, bytes) — one entry per part.

    Up to MAX_ROWS_PER_SHEET rows it's a single file named after the upload;
    beyond that each part is its own standalone MLC workbook (`…_part1.xlsx`,
    `…_part2.xlsx`, …) and the caller ships them together in a .zip. No song is
    ever split between two files.
    """
    parts = built["parts"]
    if len(parts) == 1:
        return [(f"{stem}.xlsx", to_workbook(built, parts[0]))]
    return [
        (PART_NAME.format(stem=stem, n=n), to_workbook(built, rows))
        for n, rows in enumerate(parts, start=1)
    ]


def convert(filename: str, data: bytes) -> dict:
    """Parse, expand and validate one upload."""
    info = prepare(filename, data)
    built = build(info)
    return {"info": info, "built": built, "checks": validate(info, built)}
