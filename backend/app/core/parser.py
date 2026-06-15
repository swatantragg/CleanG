"""Parse uploaded CSV/TSV and Excel bytes into (headers, list-of-row-dicts)."""
import csv
import io

from openpyxl import load_workbook

try:
    import xlrd  # legacy .xls support
except Exception:  # pragma: no cover
    xlrd = None


def _rows_from_matrix(matrix):
    # Drop fully-empty rows.
    matrix = [
        r for r in matrix
        if any(c is not None and str(c).strip() != "" for c in r)
    ]
    if not matrix:
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in matrix[0]]
    data = []
    for r in matrix[1:]:
        obj = {}
        for i, h in enumerate(headers):
            if h == "":
                continue
            v = r[i] if i < len(r) else ""
            obj[h] = "" if v is None else str(v).strip()
        data.append(obj)

    headers = [h for h in headers if h != ""]
    return headers, data


def parse_upload(filename: str, content: bytes, ext: str):
    if ext in ("csv", "tsv"):
        delimiter = "\t" if ext == "tsv" else ","
        text = content.decode("utf-8-sig", errors="replace")
        # newline="" lets csv handle \r, \n and \r\n line endings correctly
        # (otherwise files with CR endings raise "new-line character seen in
        # unquoted field").
        reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
        return _rows_from_matrix([row for row in reader])

    if ext == "xlsx":
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        matrix = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        return _rows_from_matrix(matrix)

    if ext == "xls":
        if xlrd is None:
            raise ValueError("Legacy .xls support is unavailable on the server")
        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        matrix = [
            [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        return _rows_from_matrix(matrix)

    raise ValueError("Unsupported file type")
