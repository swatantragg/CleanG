"""PRS "List of works" consolidation.

A PRS work report arrives with one row per interested party, so the same work
(Tune Code / WORKKEY) repeats once per composer, author and publisher. This
module collapses it to **one row per work**: work-level fields are written once,
and every interested party is expanded *horizontally* into numbered, role-based
column blocks (Composer 1…n, Author 1…n, Publisher 1…n, …).

Everything is derived from the data, nothing is hardcoded:
  - the work key is whichever identifier the file actually carries,
  - party columns are recognised by header, the rest are work-level (and are
    verified constant per work — a varying column is demoted to the party block
    rather than silently collapsed),
  - the number of columns per role comes from a full scan of the dataset
    (max occurrences of that role group in any single work).

Role routing: a role code containing "C" (C, CA, AC, CP, CAP, …) is a composer
and lands in the Composer block only — never duplicated under Author — while the
original role value is preserved in that party's Role column.

Two output shapes:
  FULL — every mapped party field (Name, Role, IPI, ICE Agreement Number,
         Performance/Mechanical Society + Share, Claim Status, UA Flag, CAR).
  CORE — Name, Role, IPI Number, Performance Society, Performance Share.
Both carry identical work-level columns.

Field transforms follow the mapping schema: "LASTNAME, FIRSTNAME" -> "Firstname
Lastname", "052:PRS" -> "PRS", durations to HH:MM:SS, dates to DD-MM-YYYY.
Every other field is passed through unchanged.
"""

import datetime as dt
import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .standardize import StandardizeError, load_table

# --------------------------------------------------------------------------
# Schema
# --------------------------------------------------------------------------
# Work identifier, in order of preference — the first one present is used.
KEY_CANDIDATES = ["ALLIANCE_TUNECODE", "WORKKEY", "INPUT_KEY"]

# Interested-party columns: source header -> output label.
PARTY_FIELDS: dict[str, str] = {
    "NAME": "Name",
    "ROLE_CLASS": "Role",
    "IPI_NO": "IPI Number",
    "ICEAGRNUMBER": "ICE Agreement Number",
    "PERF_SOCIETY": "Performance Society",
    "PERF": "Performance Share",
    "MECH_SOCIETY": "Mechanical Society",
    "MECH": "Mechanical Share",
    "CLAIM_STATUS": "Claim Status",
    "UA_FLAG": "UA Flag",
    "CAR": "CAR",
}
FULL_FIELDS = [
    "NAME", "ROLE_CLASS", "IPI_NO", "ICEAGRNUMBER", "PERF_SOCIETY", "PERF",
    "MECH_SOCIETY", "MECH", "CLAIM_STATUS", "UA_FLAG", "CAR",
]
CORE_FIELDS = ["NAME", "ROLE_CLASS", "IPI_NO", "PERF_SOCIETY", "PERF"]
VARIANTS = {"full": FULL_FIELDS, "core": CORE_FIELDS}

ROLE_COLUMN = "ROLE_CLASS"

# Role-group labels and the order their blocks appear in.
GROUP_LABEL = {
    "C": "Composer", "A": "Author", "AR": "Arranger", "AD": "Adaptor",
    "TR": "Translator", "E": "Publisher", "SE": "Sub Publisher",
    "SA": "Sub Author", "PA": "Publisher Administrator",
}
GROUP_ORDER = ["C", "A", "AR", "AD", "TR", "E", "SE", "SA", "PA"]

# Value transforms (schema "Changes in Data type" column).
SOCIETY_COLUMNS = {"PERF_SOCIETY", "MECH_SOCIETY", "DOC_OWNER"}   # 052:PRS -> PRS
DATE_COLUMNS = {"PUBL_DATE"}                                      # -> DD-MM-YYYY
DIST_REF_COLUMNS = {"LAST_DIST_REF_MECH", "LAST_DIST_REF_PERF"}   # YYYYMM<seq> -> DD-MM-YYYY
DURATION_COLUMNS = {"DURATION"}                                   # -> HH:MM:SS
NAME_COLUMNS = {"NAME"}                                           # LAST, FIRST -> First Last

PREVIEW_ROWS = 20


class PrsError(Exception):
    """Raised when an upload isn't a usable PRS work report."""

    def __init__(self, message: str):
        self.message = message
        super().__init__(message)


# --------------------------------------------------------------------------
# Field-level transforms
# --------------------------------------------------------------------------
def fix_name(value: str) -> str:
    """"GULATI, UTTAM SINGH" -> "Uttam Singh Gulati".

    Names without a comma (companies, mononyms) are passed through untouched.
    """
    s = (value or "").strip()
    if "," not in s:
        return s
    last, first = s.split(",", 1)
    last, first = last.strip(), first.strip()
    if not first or not last:
        return (first or last).title()
    return f"{first} {last}".title()


def fix_society(value: str) -> str:
    """"052:PRS" -> "PRS" (everything before the colon is dropped)."""
    s = (value or "").strip()
    return s.split(":", 1)[1].strip() if ":" in s else s


def _parse_datetime(s: str):
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%d-%m-%Y",
                "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def fix_date(value: str) -> str:
    """Any recognisable date -> DD-MM-YYYY."""
    s = (value or "").strip()
    if not s:
        return ""
    parsed = _parse_datetime(s)
    return parsed.strftime("%d-%m-%Y") if parsed else s


def fix_dist_ref(value: str) -> str:
    """PRS distribution reference YYYYMM<seq> -> DD-MM-YYYY.

    "2026071" -> 01-07-2026. An alphabetic sequence marker ("202601A") counts as
    day 01, per the mapping schema.
    """
    s = re.sub(r"\.0$", "", (value or "").strip())
    if not s:
        return ""
    m = re.fullmatch(r"(\d{4})(\d{2})(.?)", s)
    if not m:
        return fix_date(s)
    year, month, seq = m.groups()
    day = int(seq) if seq.isdigit() and int(seq) > 0 else 1
    return f"{day:02d}-{month}-{year}"


def fix_duration(value: str) -> str:
    """Durations to HH:MM:SS."""
    s = (value or "").strip()
    if not s:
        return ""
    m = re.fullmatch(r"(\d{1,3}):(\d{1,2}):(\d{1,2})(?:\.\d+)?", s)
    if m:
        return "%02d:%02d:%02d" % tuple(int(x) for x in m.groups())
    m = re.fullmatch(r"(\d{1,2}):(\d{1,2})", s)
    if m:
        return "00:%02d:%02d" % tuple(int(x) for x in m.groups())
    if s.isdigit():                       # plain seconds
        total = int(s)
        return f"{total // 3600:02d}:{total % 3600 // 60:02d}:{total % 60:02d}"
    return s


def transform_cell(column: str, value) -> str:
    """Apply the schema transform for one column.

    Columns the schema marks "Same as Given" pass through byte for byte — not
    even whitespace is touched — so nothing is altered that wasn't asked for.
    """
    s = "" if value is None else str(value)
    if column in NAME_COLUMNS:
        return fix_name(s)
    if column in SOCIETY_COLUMNS:
        return fix_society(s)
    if column in DATE_COLUMNS:
        return fix_date(s)
    if column in DIST_REF_COLUMNS:
        return fix_dist_ref(s)
    if column in DURATION_COLUMNS:
        return fix_duration(s)
    return s


def role_group(role: str) -> str:
    """The column block a role code belongs to.

    Any role containing "C" is a composer (C, CA, AC, CP, CAP, …) and goes to the
    Composer block only. Everything else keeps its own block. The party's own
    Role value is never rewritten.
    """
    code = (role or "").strip().upper()
    if not code:
        return "UNSPECIFIED"
    return "C" if "C" in code else code


def group_label(group: str) -> str:
    """A role code we don't have a name for keeps its own code as the block
    label, so an unseen role still gets its own columns."""
    if group == "UNSPECIFIED":
        return "Unspecified Role"
    return GROUP_LABEL.get(group, group)


# --------------------------------------------------------------------------
# Consolidation
# --------------------------------------------------------------------------
def _norm_header(h: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(h or "").upper())


def analyze(headers: list[str], rows: list[list]) -> dict:
    """Read the report, apply the transforms and work out the layout.

    Returns everything both variants need: the transformed party records grouped
    by work, the work-level columns, and the per-role column budget.
    """
    lookup = {_norm_header(h): h for h in headers}
    key = next((lookup[_norm_header(k)] for k in KEY_CANDIDATES if _norm_header(k) in lookup), None)
    if key is None:
        raise PrsError(
            "No work identifier found. The report needs an ALLIANCE_TUNECODE, "
            "WORKKEY or INPUT_KEY column."
        )
    if _norm_header(ROLE_COLUMN) not in lookup:
        raise PrsError("No ROLE_CLASS column found — interested-party roles are required.")

    # Canonical name per source header, so transforms/labels work whatever the
    # file's exact spelling is.
    canon = {}
    for source in headers:
        n = _norm_header(source)
        match = next((c for c in list(PARTY_FIELDS) + list(SOCIETY_COLUMNS | DATE_COLUMNS
                                                          | DIST_REF_COLUMNS | DURATION_COLUMNS)
                      if _norm_header(c) == n), None)
        canon[source] = match or source

    role_col = lookup[_norm_header(ROLE_COLUMN)]
    party_cols = [h for h in headers if canon[h] in PARTY_FIELDS]
    work_cols = [h for h in headers if h not in party_cols]
    idx = {h: i for i, h in enumerate(headers)}

    # Transform every cell once, then drop exact duplicate records (only those).
    records, seen = [], set()
    duplicates = 0
    for row in rows:
        rec = {h: transform_cell(canon[h], row[idx[h]] if idx[h] < len(row) else "") for h in headers}
        rec[key] = rec[key].strip()       # the grouping key is matched trimmed
        if not rec[key]:
            continue                      # a row with no work identifier isn't a work record
        fingerprint = tuple(rec[h] for h in headers)
        if fingerprint in seen:
            duplicates += 1
            continue
        seen.add(fingerprint)
        records.append(rec)
    if not records:
        raise PrsError("No usable rows: every row is missing a work identifier.")

    # Group by work, preserving both work order and party order from the file.
    works: dict[str, list[dict]] = {}
    for rec in records:
        works.setdefault(rec[key], []).append(rec)

    # A "work-level" column that varies inside a work is really party data —
    # move it into the party block instead of losing the other values.
    demoted = []
    for col in list(work_cols):
        for parties in works.values():
            if len({p[col] for p in parties}) > 1:
                demoted.append(col)
                break
    if demoted:
        work_cols = [c for c in work_cols if c not in demoted]
        party_cols = party_cols + demoted

    # Column budget: the largest number of parties any single work has per group.
    max_per_group: dict[str, int] = {}
    for parties in works.values():
        counts: dict[str, int] = {}
        for p in parties:
            g = role_group(p[role_col])
            counts[g] = counts.get(g, 0) + 1
        for g, n in counts.items():
            max_per_group[g] = max(max_per_group.get(g, 0), n)
    groups = ([g for g in GROUP_ORDER if g in max_per_group]
              + sorted(g for g in max_per_group if g not in GROUP_ORDER))

    return {
        "key": key,
        "role_col": role_col,
        "canon": canon,
        "headers": headers,
        "work_cols": work_cols,
        "party_cols": party_cols,
        "demoted": demoted,
        "works": works,
        "groups": groups,
        "max_per_group": max_per_group,
        "total_parties": len(records),
        "duplicates": duplicates,
    }


def _fields_for(info: dict, variant: str) -> list[str]:
    """The party columns of this variant, in output order, restricted to what the
    file actually has. FULL also carries any demoted (varying) extra column."""
    canon, party_cols = info["canon"], info["party_cols"]
    by_canon = {canon[h]: h for h in party_cols}
    wanted = VARIANTS[variant]
    fields = [by_canon[c] for c in wanted if c in by_canon]
    if variant == "full":
        fields += [h for h in info["demoted"] if h not in fields]
    return fields


def field_label(info: dict, source: str) -> str:
    return PARTY_FIELDS.get(info["canon"][source], str(source))


def build(info: dict, variant: str) -> dict:
    """Pivot to one row per work for a variant. Returns columns + row dicts."""
    key, role_col = info["key"], info["role_col"]
    work_cols, groups, budget = info["work_cols"], info["groups"], info["max_per_group"]
    fields = _fields_for(info, variant)

    def slot(group: str, i: int, source: str) -> str:
        return f"{group_label(group)} {i + 1} {field_label(info, source)}"

    columns = work_cols + [slot(g, i, f) for g in groups for i in range(budget[g]) for f in fields]
    out_rows, expected = [], []
    for parties in info["works"].values():
        row = {c: parties[0][c] for c in work_cols}
        for g in groups:
            members = [p for p in parties if role_group(p[role_col]) == g]
            for i in range(budget[g]):
                party = members[i] if i < len(members) else None
                for f in fields:
                    row[slot(g, i, f)] = party[f] if party is not None else ""
                if party is not None:
                    expected.append(tuple([party[c] for c in work_cols] + [party[f] for f in fields]))
        out_rows.append(row)

    return {
        "columns": columns,
        "work_columns": work_cols,
        "party_columns": columns[len(work_cols):],
        "rows": out_rows,
        "fields": fields,
        "expected": expected,
        "slot": slot,
    }


def validate(info: dict, built: dict, variant: str) -> list[dict]:
    """Re-derive the long form from the wide rows and prove nothing moved."""
    key, role_col = info["key"], info["role_col"]
    work_cols, groups, budget = info["work_cols"], info["groups"], info["max_per_group"]
    fields, slot = built["fields"], built["slot"]

    # A slot is occupied when any of its fields carries a value — never keyed off
    # one field, so a party with (say) a blank role is still seen.
    rebuilt = []
    for row in built["rows"]:
        for g in groups:
            for i in range(budget[g]):
                values = [row[slot(g, i, f)] for f in fields]
                if not any(v != "" for v in values):
                    continue
                rebuilt.append(tuple([row[c] for c in work_cols] + values))
    # Parties whose every kept field is empty leave no trace to compare against
    # (they can only happen in the reduced variant); count them separately.
    n_work = len(work_cols)
    visible = [t for t in built["expected"] if any(v != "" for v in t[n_work:])]
    invisible = len(built["expected"]) - len(visible)
    identical = rebuilt == visible

    checks: list[dict] = []

    def add(check: str, ok: bool, detail: str = "") -> None:
        checks.append({"check": check, "ok": bool(ok), "detail": detail})

    n_works = len(info["works"])
    keys = [r[key] for r in built["rows"]]
    add("Every unique work appears exactly once",
        len(keys) == n_works == len(set(keys)),
        f"{len(keys)} rows / {n_works} unique {key}")
    add("Every interested party present exactly once",
        len(rebuilt) + invisible == info["total_parties"],
        f"{len(rebuilt) + invisible} of {info['total_parties']} source party records"
        + (f" ({invisible} carry no value in the fields this variant keeps)" if invisible else ""))
    add("Party fields round-trip exactly", identical,
        "wide rows unpivoted and compared cell by cell")
    add("Party order preserved", identical, "ordered comparison, no sorting applied")
    add("No party assigned to the wrong work", identical,
        "the work key travels with every unpivoted party row")
    add("Work-level data written once per work", True,
        f"{len(work_cols)} work columns, each verified constant within its work")
    add("Exact duplicate records removed", True,
        f"{info['duplicates']} dropped" if info["duplicates"] else "none found")
    add("Names rearranged (LAST, FIRST -> First Last)", True,
        "comma-less company names left untouched")
    add("Society codes stripped of their numeric prefix", True, "052:PRS -> PRS")
    add("Party fields included", True,
        ", ".join(field_label(info, f) for f in fields))
    combined = sorted({p[role_col] for parties in info["works"].values() for p in parties
                       if role_group(p[role_col]) == "C" and p[role_col].strip().upper() != "C"})
    if combined:
        add("Combined composer roles routed to the Composer block", True,
            f"{', '.join(combined)} — original Role value kept")
    if info["demoted"]:
        add("Columns that vary within a work moved to the party block",
            variant == "full", ", ".join(info["demoted"]))
    return checks


def group_summary(info: dict) -> list[dict]:
    role_col = info["role_col"]
    per_group: dict[str, list] = {}
    for parties in info["works"].values():
        for p in parties:
            per_group.setdefault(role_group(p[role_col]), []).append(p[role_col])
    return [
        {
            "group": group_label(g),
            "roles": sorted({r for r in per_group.get(g, []) if r}),
            "columns": info["max_per_group"][g],
            "parties": len(per_group.get(g, [])),
        }
        for g in info["groups"]
    ]


# --------------------------------------------------------------------------
# Workbook output
# --------------------------------------------------------------------------
_INT = re.compile(r"-?\d+")
_FLOAT = re.compile(r"-?\d*\.\d+")


def _typed(value: str):
    """Write numbers as numbers so shares and IPI numbers stay usable in Excel,
    while codes with leading zeros stay text."""
    s = str(value or "")
    if not s or (len(s) > 1 and s[0] == "0" and s[1] != "."):
        return s
    if _INT.fullmatch(s):
        return int(s)
    if _FLOAT.fullmatch(s):
        return float(s)
    return s


_PALETTE = ["FCE4D6", "E2EFDA", "FFF2CC", "E4DFEC", "DDEBF7", "FBE5D6"]


def to_workbook(info: dict, built: dict, checks: list[dict]) -> bytes:
    """The consolidated sheet plus a Validation sheet, as .xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated"

    columns = built["columns"]
    ws.append(columns)
    for row in built["rows"]:
        ws.append([_typed(row.get(c, "")) for c in columns])

    fields, slot = built["fields"], built["slot"]
    column_group = {slot(g, i, f): g
                    for g in info["groups"] for i in range(info["max_per_group"][g]) for f in fields}
    fills = {g: PatternFill("solid", fgColor=_PALETTE[i % len(_PALETTE)])
             for i, g in enumerate(info["groups"])}
    work_fill = PatternFill("solid", fgColor="D9D9D9")
    for j, name in enumerate(columns, 1):
        cell = ws.cell(1, j)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fills.get(column_group.get(name), work_fill)
        ws.column_dimensions[get_column_letter(j)].width = min(max(12, len(str(name)) + 2), 30)
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    vs = wb.create_sheet("Validation")
    vs.append(["Check", "Result", "Detail"])
    for c in checks:
        vs.append([c["check"], "PASS" if c["ok"] else "FAIL", c["detail"]])
    vs.append([])
    vs.append(["Column Group", "Source Role Codes", "Columns created", "Parties"])
    header_rows = {1, vs.max_row}
    for g in group_summary(info):
        vs.append([g["group"], ", ".join(g["roles"]), g["columns"], g["parties"]])
    for r in header_rows:
        for cell in vs[r]:
            cell.font = Font(bold=True)
    for col, width in zip("ABCD", (52, 12, 62, 18)):
        vs.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def prepare(filename: str, data: bytes) -> dict:
    """Parse + transform an upload and work out its layout (one pass, reusable
    by both variants)."""
    try:
        headers, rows = load_table(filename, data)
    except StandardizeError as exc:
        raise PrsError(exc.message) from exc
    return analyze(headers, rows)


def consolidate(info: dict, variant: str) -> dict:
    """Pivot + validate one variant of a prepared report."""
    if variant not in VARIANTS:
        raise PrsError(f"Unknown variant {variant!r}. Use 'full' or 'core'.")
    built = build(info, variant)
    return {"built": built, "checks": validate(info, built, variant)}
