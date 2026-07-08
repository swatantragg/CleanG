import io
import json
from collections import Counter

from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session, defer

from ..config import get_settings
from ..core.limiter import limiter

from ..core.cleaning import (
    FIELD_TYPES,
    FIX_LABELS,
    NAME_SEP,
    NAME_TYPES,
    TAG_LABELS,
    CleanRow,
    _dedup_norm,
    clean_dataset,
    mark_duplicates,
    revalidate,
)
from ..core.http import content_disposition, safe_filename
from ..core.master_store import find_conflicts, upsert_master_records
from ..database import get_db
from ..deps import get_current_user
from ..models import (
    MASTER_COLUMN_TO_ATTR,
    ActivityLog,
    Branch,
    FileStatus,
    UploadedFile,
    User,
    UserRole,
)
from ..schemas import (
    BulkFix,
    CleanRowOut,
    CleanSummary,
    ColumnFill,
    ColumnProfile,
    CommitRequest,
    CommitResult,
    ConflictsResult,
    DataProfile,
    RemapPreview,
    ReviewOut,
    RowEdit,
    RowsAccept,
    RowsBatchEdit,
    RowsRevert,
    TagGroup,
    UniqueValue,
    UniqueValuesOut,
    ValueRemap,
)

# The dataset heatmap ribbon is capped at this many pixels; larger files are
# downsampled (each pixel takes the worst status of the rows it represents) so
# the payload stays small no matter how big the upload is.
_MAX_STRIP = 1800


def _grade(score: int) -> str:
    table = [
        (97, "A+"), (93, "A"), (90, "A-"),
        (87, "B+"), (83, "B"), (80, "B-"),
        (77, "C+"), (73, "C"), (70, "C-"),
        (60, "D"),
    ]
    for cutoff, letter in table:
        if score >= cutoff:
            return letter
    return "F"

settings = get_settings()
router = APIRouter(tags=["clean"])


def _get_file_or_404(file_id: int, user: User, db: Session) -> UploadedFile:
    # Defer the big `data` blob: review endpoints only touch it on a cache miss
    # (via build_rows), so a warm cache serves the page with no large read.
    f = db.scalars(
        select(UploadedFile)
        .where(UploadedFile.id == file_id)
        .options(defer(UploadedFile.data))
    ).first()
    if f is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "File not found")
    branch = db.get(Branch, f.branch_id)
    if user.role != UserRole.admin and branch.owner_id != user.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Not your branch")
    return f


def _active_columns(f: UploadedFile) -> list[str]:
    active = {m["master_column"] for m in f.mapping if m.get("input_header")}
    # Columns given a constant fill are output columns even with no input source.
    active |= set(f.constants or {})
    # Lead Artist is auto-derived from Singer/Composer/Lyricist, so it's a real
    # output column whenever any contributing credit (or Lead Artist) is present.
    if active & {"Singer", "Composer", "Lyricist", "Lead Artist"}:
        active.add("Lead Artist")
    # Canonical master columns first, in schema order so derived/filled columns
    # slot into their natural position; user-added custom columns (not part of the
    # fixed schema, e.g. "Mood") follow, in the order their mapping rows appear.
    canonical = [c for c in MASTER_COLUMN_TO_ATTR if c in active]
    custom = [
        m["master_column"]
        for m in f.mapping
        if m["master_column"] in active
        and m["master_column"] not in MASTER_COLUMN_TO_ATTR
    ]
    seen = set(canonical)
    return canonical + [c for c in custom if not (c in seen or seen.add(c))]


def _is_name_column(col: str) -> bool:
    """Name fields (Singer, Composer, ...) hold pipe-separated people, so their
    distinct/unique values are counted per individual name, not per whole cell."""
    return FIELD_TYPES.get(col, "text") in NAME_TYPES


def _value_pieces(col: str, value: str) -> list[str]:
    """The distinct units of a cell: each piped name for name fields, else the
    whole trimmed value. Empty cells contribute nothing."""
    v = (value or "").strip()
    if not v:
        return []
    if _is_name_column(col):
        return [p.strip() for p in v.split(NAME_SEP) if p.strip()]
    return [v]


# --------------------------------------------------------------------------
# In-memory cleaning
#
# Cleaning is a pure function of the file's raw `data` + `mapping`, with the
# reviewer's `corrections`/`dropped` overlaid on top. We never persist the
# cleaned rows — they're recomputed on demand (milliseconds) and memoised
# per-process so repeated review requests don't re-clean from scratch.
# --------------------------------------------------------------------------
# Per-process cache: keyed by file id, holds the cleaned rows AND the derived
# summary/profile so repeated review/paging requests (and the clean -> review
# hand-off) never re-clean or re-aggregate. All three are pure functions of the
# file's signature, so a single signature check keeps them consistent.
_CACHE: dict[int, dict] = {}
_CACHE_LIMIT = 16


def _signature(f: UploadedFile) -> str:
    # `data` is immutable after upload, so only the overlay + mapping vary.
    return json.dumps(
        [
            f.mapping, f.corrections or {}, f.dropped or [],
            f.accepted or [], f.constants or {},
        ],
        sort_keys=True, default=str,
    )


def _mark_corrected(
    base_values: dict, cleaned: dict, issues: list[dict], override: dict
) -> list[dict]:
    """Flag every cell a human changed (merge / inline edit / bulk set) so the
    Review grid highlights it as "Manually corrected".

    A cell is corrected when its final value differs from the value cleaning
    produced before the overlay. The marker carries the pre-correction value as
    `original`, so hovering shows what it was. Cells still in error keep their
    error flag (it already highlights and explains the problem); any cosmetic
    auto-fix flag on a corrected cell is superseded by this stronger marker.
    """
    edited = {
        col for col in override
        if (cleaned.get(col) or "") != (base_values.get(col) or "")
    }
    if not edited:
        return issues
    err_cols = {i["column"] for i in issues if i["action"] == "error"}
    kept = [i for i in issues if i["column"] not in edited or i["action"] == "error"]
    for col in edited:
        if col in err_cols:
            continue
        kept.append({
            "column": col,
            "action": "fixed",
            "tag": "corrected",
            "message": "Manually corrected in review.",
            "value": cleaned.get(col, ""),
            "original": base_values.get(col, ""),
            "cosmetic": False,
        })
    return kept


def _entry(f: UploadedFile) -> dict:
    """The cache entry for a file (rows + memoised summary/profile), built lazily."""
    sig = _signature(f)
    hit = _CACHE.get(f.id)
    if hit and hit["sig"] == sig:
        return hit

    base = clean_dataset(f.data, f.headers, f.mapping, f.constants or {})
    corrections = f.corrections or {}
    dropped = set(f.dropped or [])
    accepted = set(f.accepted or [])

    rows: list[CleanRow] = []
    for r in base:
        if r.index in dropped:
            continue
        override = corrections.get(str(r.index))
        if override:
            cleaned, issues = revalidate({**r.values, **override})
            issues = _mark_corrected(r.values, cleaned, issues, override)
            r = CleanRow(index=r.index, values=cleaned, issues=issues)
        rows.append(r)

    mark_duplicates(rows)

    # "Keep as-is": clear error flags on accepted rows (values stay unchanged),
    # so the reviewer's deliberate keep moves the row into the clean set.
    if accepted:
        for r in rows:
            if r.index in accepted:
                r.issues = [i for i in r.issues if i["action"] != "error"]

    if len(_CACHE) >= _CACHE_LIMIT:
        _CACHE.pop(next(iter(_CACHE)))
    # `base_values` is the pre-overlay cleaned value dict for EVERY data row (keyed
    # by original index). It lets a later single/batch edit or artist merge re-clean
    # only the changed rows (see _edit_in_place) instead of re-cleaning the whole
    # file — captured before mark_duplicates, which only touches issues, not values.
    entry = {
        "sig": sig,
        "rows": rows,
        "summary": None,
        "profile": None,
        "base_values": {r.index: r.values for r in base},
    }
    _CACHE[f.id] = entry
    return entry


def _warm_base_values(f: UploadedFile) -> dict | None:
    """The cached per-row base values if the file's clean cache is warm, else None.

    Warm access means an edit can validate its row index and recompute in place
    without loading the (large, possibly remote) `data` blob at all."""
    e = _CACHE.get(f.id)
    if e and e.get("base_values") is not None:
        return e["base_values"]
    return None


def _edit_in_place(f: UploadedFile, changed: set[int]) -> bool:
    """Recompute only the `changed` rows in the warm cache after their corrections
    were updated (written, or removed by a revert), instead of re-cleaning the
    entire file.

    Returns False if the cache is cold or in an unexpected state, so the caller
    falls back to a full invalidate + rebuild — always correct, just slower. This
    is the hot path for inline edits and artist merges on large files: it avoids
    both the ~1s whole-file re-clean and the multi-MB data-blob reload from the
    (remote) database that a full rebuild would trigger on the next request.
    """
    if not changed:
        return True  # nothing to recompute; the warm cache is already consistent
    entry = _CACHE.get(f.id)
    if not entry or entry.get("base_values") is None:
        return False
    base_values: dict = entry["base_values"]
    corrections = f.corrections or {}
    dropped = set(f.dropped or [])
    accepted = set(f.accepted or [])
    rows: list[CleanRow] = entry["rows"]
    pos = {r.index: k for k, r in enumerate(rows)}

    for idx in changed:
        bv = base_values.get(idx)
        # A revert clears the row's override, so a missing one is expected there;
        # the row simply re-cleans back to its pre-review values. A row that's
        # absent or dropped is unexpected -> bail to the safe full-rebuild path.
        if bv is None or idx in dropped or idx not in pos:
            return False
        override = corrections.get(str(idx)) or {}
        cleaned, issues = revalidate({**bv, **override})
        issues = _mark_corrected(bv, cleaned, issues, override)
        rows[pos[idx]] = CleanRow(index=idx, values=cleaned, issues=issues)

    # An edit can change cross-row duplicate flags; re-run the cheap (no-DB) pass.
    mark_duplicates(rows)
    # Preserve the "keep as-is" decisions (accepted rows keep their errors cleared).
    if accepted:
        for r in rows:
            if r.index in accepted:
                r.issues = [i for i in r.issues if i["action"] != "error"]
    entry["sig"] = _signature(f)
    entry["summary"] = None
    entry["profile"] = None
    return True


def build_rows(f: UploadedFile) -> list[CleanRow]:
    """The cleaned, corrected, duplicate-checked rows for a file (cached)."""
    return _entry(f)["rows"]


def _get_summary(f: UploadedFile) -> "CleanSummary":
    e = _entry(f)
    if e["summary"] is None:
        e["summary"] = _summary(f, e["rows"])
    return e["summary"]


def _get_profile(f: UploadedFile) -> "DataProfile":
    e = _entry(f)
    if e["profile"] is None:
        e["profile"] = _build_profile(_active_columns(f), e["rows"])
    return e["profile"]


def _invalidate(file_id: int) -> None:
    _CACHE.pop(file_id, None)


def _accept_in_place(f: UploadedFile) -> None:
    """Apply the file's accepted-set to the cached rows WITHOUT re-cleaning.

    Accepting a row only clears its error flags (the values are unchanged) — which
    is exactly what `_entry` does for accepted rows. So instead of invalidating the
    cache and re-cleaning every row (expensive on large files), we mutate the hot
    cache in place: clear errors on the accepted rows, refresh the signature, and
    drop the memoised summary/profile so they recompute cheaply (no cleaning) on
    next access. Falls back to a normal lazy rebuild if the cache is cold."""
    entry = _CACHE.get(f.id)
    if not entry:
        return
    accepted = set(f.accepted or [])
    for r in entry["rows"]:
        if r.index in accepted:
            r.issues = [i for i in r.issues if i["action"] != "error"]
    entry["sig"] = _signature(f)
    entry["summary"] = None
    entry["profile"] = None


def _row_key(key: object) -> int | None:
    """A `corrections` dict key as a row index, or None when it isn't one."""
    try:
        return int(key)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _manual_kinds(f: UploadedFile) -> dict[int, str]:
    """row_index -> how a human moved that row into the clean set.

    "edited" — the reviewer changed cell values (inline edit, bulk set, or an
    artist merge), stored as a correction.
    "kept"   — the reviewer hit "keep as-is", clearing the row's error flags
    without touching any value.

    A row that was both edited and accepted reports "edited" (the stronger,
    more informative signal). Rows absent from this map were cleaned entirely by
    the tool — the "Auto cleaned" tab.
    """
    kinds: dict[int, str] = {i: "kept" for i in (f.accepted or [])}
    for key in (f.corrections or {}):
        idx = _row_key(key)
        if idx is not None:
            kinds[idx] = "edited"
    return kinds


def _row_out(r: CleanRow, manual_kind: str | None = None) -> CleanRowOut:
    return CleanRowOut(
        row_index=r.index,
        status=r.status,
        values=r.values,
        issues=r.issues,
        manual_kind=manual_kind,
    )


def _review_payload(
    f: UploadedFile,
    view: str,
    tag: str | None,
    page: int,
    page_size: int,
    include_profile: bool,
    tags: list[str] | None = None,
    sort: str | None = None,
    direction: str = "asc",
    contains_col: str | None = None,
    contains_val: str | None = None,
) -> "ReviewOut":
    """Build the full Review payload (summary + profile + page of rows). Shared by
    GET /review and the edit/accept mutations so they each return in one trip.

    Optional `tags` (match ANY), `contains_col`/`contains_val` (value filter) and
    `sort`/`direction` are applied server-side so they hold across pages."""
    rows = build_rows(f)
    manual = _manual_kinds(f)
    filtered = _filter_rows(rows, view, tag, tags, contains_col, contains_val, manual)
    filtered = _sort_rows(filtered, sort, direction)
    total = len(filtered)
    page = max(0, page)
    page_rows = filtered[page * page_size : page * page_size + page_size]
    return ReviewOut(
        summary=_get_summary(f),
        profile=_get_profile(f) if include_profile else None,
        rows=[_row_out(r, manual.get(r.index)) for r in page_rows],
        total=total,
        page=page,
        page_size=page_size,
    )


def _summary(f: UploadedFile, rows: list[CleanRow]) -> CleanSummary:
    tags: Counter = Counter()
    fix_tags: Counter = Counter()
    manual = _manual_kinds(f)
    auto_fixed = 0
    clean = 0
    manual_clean = 0
    for r in rows:
        if r.status == "clean":
            clean += 1
            if r.index in manual:
                manual_clean += 1
        for i in r.issues:
            if i["action"] == "error":
                tags[i["tag"]] += 1
            elif i["action"] == "fixed":
                auto_fixed += 1
                fix_tags[i.get("tag") or "trimmed"] += 1
    return CleanSummary(
        total=len(rows),
        clean=clean,
        auto_clean=clean - manual_clean,
        manual_clean=manual_clean,
        errors=len(rows) - clean,
        auto_fixed=auto_fixed,
        tags=[
            TagGroup(tag=t, label=TAG_LABELS.get(t, t), count=c)
            for t, c in tags.most_common()
        ],
        fix_tags=[
            TagGroup(tag=t, label=FIX_LABELS.get(t, t), count=c)
            for t, c in fix_tags.most_common()
        ],
        columns=_active_columns(f),
    )


@router.post("/api/files/{file_id}/clean", response_model=CleanSummary)
def run_clean(
    file_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Mark the file as cleaned and return a fresh summary.

    Cleaning itself is computed in memory — nothing is written to the database
    here beyond flipping the file's status.
    """
    f = _get_file_or_404(file_id, user, db)
    if not _active_columns(f):
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "Map at least one column before cleaning.",
        )
    # Cleaning is a pure function of the file's signature, so a warm cache is
    # reused (no forced re-clean). The signature changes if the mapping did.
    summary = _get_summary(f)
    # Warm the quality profile too, while the "Run cleaning" progress bar is up,
    # so landing on Review serves rows + summary + profile from a hot cache.
    _get_profile(f)
    if f.status not in (FileStatus.cleaned, FileStatus.committed):
        f.status = FileStatus.cleaned
        db.commit()
    return summary


@router.get("/api/files/{file_id}/clean/summary", response_model=CleanSummary)
def clean_summary(
    file_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    f = _get_file_or_404(file_id, user, db)
    return _get_summary(f)


def _build_profile(columns: list[str], rows: list[CleanRow]) -> DataProfile:
    """Profile the cleaned dataset: an overall quality score, a per-row health
    ribbon, and per-column completeness/quality stats. Pure data engineering —
    everything is derived from the in-memory cleaned rows."""
    # Per-column accumulators.
    filled = {c: 0 for c in columns}
    fixed = {c: 0 for c in columns}
    normalized = {c: 0 for c in columns}
    errors = {c: 0 for c in columns}
    distinct: dict[str, set] = {c: set() for c in columns}
    # Pipe-aware distinct: name fields count each individual name (Singer 1 |
    # Singer 2 -> 2), so the per-column "unique" badge reflects real people.
    distinct_split: dict[str, set] = {c: set() for c in columns}
    counters: dict[str, Counter] = {c: Counter() for c in columns}

    # Per-row worst status (0 ok, 1 fixed, 2 error) for the heatmap ribbon.
    # Cosmetic-only normalization stays green — only meaningful corrections amber.
    worst: list[int] = []
    clean_rows = 0
    fixed_cells = 0
    normalized_cells = 0
    error_cells = 0

    for r in rows:
        if r.status == "clean":
            clean_rows += 1
        row_worst = 0
        for c in columns:
            v = (r.values.get(c) or "").strip()
            if v:
                filled[c] += 1
                distinct[c].add(v)
                counters[c][v] += 1
                distinct_split[c].update(_value_pieces(c, v))
        for issue in r.issues:
            col = issue.get("column")
            if col not in filled:
                continue
            if issue["action"] == "error":
                errors[col] += 1
                error_cells += 1
                row_worst = 2
            elif issue["action"] == "fixed":
                if issue.get("cosmetic"):
                    normalized[col] += 1
                    normalized_cells += 1
                else:
                    fixed[col] += 1
                    fixed_cells += 1
                    row_worst = max(row_worst, 1)
        worst.append(row_worst)

    total_rows = len(rows)
    n_cols = max(1, len(columns))
    total_cells = total_rows * n_cols
    filled_cells = sum(filled.values())
    blank_cells = total_cells - filled_cells
    clean_cells = max(0, total_cells - error_cells - fixed_cells)  # normalized counts as clean

    # Quality score blends cell correctness, row cleanliness and completeness.
    if total_cells == 0:
        score = 0
    else:
        cell_quality = 1 - error_cells / total_cells
        row_quality = clean_rows / total_rows if total_rows else 1
        completeness = filled_cells / total_cells
        score = round(100 * (0.5 * cell_quality + 0.3 * row_quality + 0.2 * completeness))
        score = max(0, min(100, score))

    # Downsample the ribbon if the file is large.
    scale = 1
    strip = worst
    if len(worst) > _MAX_STRIP:
        scale = (len(worst) + _MAX_STRIP - 1) // _MAX_STRIP
        strip = [
            max(worst[i : i + scale]) for i in range(0, len(worst), scale)
        ]

    col_profiles = [
        ColumnProfile(
            name=c,
            filled=filled[c],
            blank=total_rows - filled[c],
            distinct=len(distinct[c]),
            unique=len(distinct_split[c]),
            fixed=fixed[c],
            normalized=normalized[c],
            errors=errors[c],
            completeness=round(filled[c] / total_rows, 4) if total_rows else 0.0,
            top_values=[[v, n] for v, n in counters[c].most_common(5)],
        )
        for c in columns
    ]

    return DataProfile(
        score=score,
        grade=_grade(score),
        total_rows=total_rows,
        clean_rows=clean_rows,
        total_cells=total_cells,
        clean_cells=clean_cells,
        fixed_cells=fixed_cells,
        normalized_cells=normalized_cells,
        error_cells=error_cells,
        blank_cells=blank_cells,
        row_strip=strip,
        strip_scale=scale,
        columns=col_profiles,
    )


def _filter_rows(
    rows: list[CleanRow],
    view: str | None,
    tag: str | None,
    tags: list[str] | None = None,
    contains_col: str | None = None,
    contains_val: str | None = None,
    manual: dict[int, str] | None = None,
) -> list[CleanRow]:
    """Narrow the cleaned rows to one Review tab, then apply the tag/value filters.

    Views: all · error (needs review) · clean (both clean tabs) ·
    auto_clean (clean with no human touch) · manual_clean (a reviewer edited the
    row or kept it as-is, sending it into the clean set)."""
    if view in ("clean", "error"):
        rows = [r for r in rows if r.status == view]
    elif view in ("auto_clean", "manual_clean"):
        touched = manual or {}
        want_manual = view == "manual_clean"
        rows = [
            r for r in rows
            if r.status == "clean" and ((r.index in touched) == want_manual)
        ]
    if tag:
        rows = [r for r in rows if any(i.get("tag") == tag for i in r.issues)]
    if tags:
        wanted = set(tags)
        rows = [r for r in rows if any(i.get("tag") in wanted for i in r.issues)]
    if contains_col and contains_val:
        target = contains_val.strip().lower()
        rows = [
            r for r in rows
            if any(p.lower() == target for p in _value_pieces(contains_col, r.values.get(contains_col, "")))
            or target in (r.values.get(contains_col, "") or "").lower()
        ]
    return rows


def _sort_rows(rows: list[CleanRow], col: str | None, direction: str) -> list[CleanRow]:
    """Stable, case-insensitive sort by one column's value. No-op if no column."""
    if not col:
        return rows
    return sorted(
        rows,
        key=lambda r: (r.values.get(col, "") or "").strip().lower(),
        reverse=(direction == "desc"),
    )


def _split_csv(value: str | None) -> list[str]:
    """Split a comma-joined query param into a clean list (drops blanks)."""
    if not value:
        return []
    return [p.strip() for p in value.split(",") if p.strip()]


@router.get("/api/files/{file_id}/clean/profile", response_model=DataProfile)
def clean_profile(
    file_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    f = _get_file_or_404(file_id, user, db)
    return _get_profile(f)


@router.get("/api/files/{file_id}/review", response_model=ReviewOut)
def review(
    file_id: int,
    view: str = "all",
    tag: str | None = None,
    tags: str | None = None,
    sort: str | None = None,
    dir: str = "asc",
    contains_col: str | None = None,
    contains_val: str | None = None,
    page: int = 0,
    page_size: int = 50,
    include_profile: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Everything the Review screen needs in ONE request: summary, quality
    profile, and the requested page of rows (filtered by view/tag).

    `tags` is a comma-joined list of cleaning/error tags (match ANY);
    `contains_col`/`contains_val` filter by a column value; `sort`/`dir` order
    the rows. All apply server-side so they hold across pages.

    Cleaning is computed once and memoised, so this is a single fast call no
    matter how the user pages or filters — and the big row blob is read at most
    once (on the first cache miss), never on a warm page."""
    f = _get_file_or_404(file_id, user, db)
    return _review_payload(
        f, view, tag, page, page_size, include_profile,
        tags=_split_csv(tags), sort=sort, direction=dir,
        contains_col=contains_col, contains_val=contains_val,
    )


@router.get("/api/files/{file_id}/clean/rows", response_model=list[CleanRowOut])
def clean_rows(
    file_id: int,
    status_filter: str | None = None,
    tag: str | None = None,
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """List cleaned rows, optionally filtered by status (clean|error) or error tag.

    All filtering/paging happens in memory over the cached cleaned rows, so it's
    instant regardless of file size.
    """
    f = _get_file_or_404(file_id, user, db)
    rows = _filter_rows(build_rows(f), status_filter, tag)
    return [_row_out(r) for r in rows[offset : offset + limit]]


_UNIQUE_CAP = 1000  # bound the side-panel payload for very high-cardinality columns


@router.get("/api/files/{file_id}/columns/unique", response_model=UniqueValuesOut)
def column_unique_values(
    file_id: int,
    column: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """The distinct values of one column (most-common first), pipe-split for name
    fields so each individual singer/composer is counted separately. Powers the
    per-column "Show unique values" side panel. Runs over the cached rows."""
    f = _get_file_or_404(file_id, user, db)
    if column not in _active_columns(f):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Column not found")
    counts: Counter = Counter()
    for r in build_rows(f):
        for piece in _value_pieces(column, r.values.get(column, "")):
            counts[piece] += 1
    return UniqueValuesOut(
        column=column,
        total_distinct=len(counts),
        values=[UniqueValue(value=v, count=n) for v, n in counts.most_common(_UNIQUE_CAP)],
    )


def _remap_cell(col: str, value: str, alias: dict[str, str]) -> str:
    """Rewrite one cell, replacing each matching piece via `alias` (keyed by the
    normalized piece). Name fields are split on NAME_SEP, each piece remapped,
    then re-joined de-duped with order preserved (so "Shreya Ghosal | Arijit" ->
    "Shreya Ghoshal | Arijit", and a cell holding both variants collapses to one).
    Non-name columns match/replace the whole value."""
    pieces = _value_pieces(col, value)
    if not pieces:
        return value
    out: list[str] = []
    for p in pieces:
        repl = alias.get(_dedup_norm(p), p)
        # The canonical value may itself be multi-name ("A | B"); split it too.
        for sub in (_value_pieces(col, repl) or [repl]):
            if sub not in out:
                out.append(sub)
    if _is_name_column(col):
        return NAME_SEP.join(out)
    return out[0] if out else ""


def _remap_alias(payload: ValueRemap) -> dict[str, str]:
    """The {normalized variant -> canonical} map for a remap request, dropping
    blanks and any variant that already equals the canonical value."""
    to = payload.to.strip()
    return {
        _dedup_norm(v): to
        for v in payload.from_values
        if _dedup_norm(v) and _dedup_norm(v) != _dedup_norm(to)
    }


@router.post(
    "/api/files/{file_id}/columns/remap/preview", response_model=RemapPreview
)
def remap_preview(
    file_id: int,
    column: str,
    payload: ValueRemap,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Dry run: how many rows a value-merge would rewrite — so the reviewer sees
    the blast radius and confirms before anything changes."""
    f = _get_file_or_404(file_id, user, db)
    if column not in _active_columns(f):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Column not found")
    alias = _remap_alias(payload)
    if not alias:
        return RemapPreview(affected_rows=0)
    affected = sum(
        1
        for r in build_rows(f)
        if _remap_cell(column, r.values.get(column, ""), alias) != r.values.get(column, "")
    )
    return RemapPreview(affected_rows=affected)


@router.post("/api/files/{file_id}/columns/remap", response_model=ReviewOut)
def remap_column_value(
    file_id: int,
    column: str,
    payload: ValueRemap,
    view: str = "all",
    tag: str | None = None,
    tags: str | None = None,
    sort: str | None = None,
    dir: str = "asc",
    contains_col: str | None = None,
    contains_val: str | None = None,
    page: int = 0,
    page_size: int = 50,
    include_profile: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Human-confirmed value merge: replace every `from_values` variant with `to`
    across the column. Piece-aware for pipe-separated name fields. Each rewrite is
    stored as a row correction — auditable, reversible (remap back), and picked up
    by review / export / commit just like an inline edit. Returns the refreshed
    Review payload so the grid updates in one round-trip."""
    f = _get_file_or_404(file_id, user, db)
    if column not in _active_columns(f):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Column not found")
    alias = _remap_alias(payload)
    if not alias:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Nothing to remap.")

    corrections = dict(f.corrections or {})
    changed: set[int] = set()
    for r in build_rows(f):
        cur = r.values.get(column, "")
        new = _remap_cell(column, cur, alias)
        if new != cur:
            override = dict(corrections.get(str(r.index), {}))
            override[column] = new
            corrections[str(r.index)] = override
            changed.add(r.index)
    f.corrections = corrections
    db.commit()
    # Recompute only the rows the merge rewrote; full rebuild only if cache is cold.
    if not _edit_in_place(f, changed):
        _invalidate(file_id)
    return _review_payload(
        f, view, tag, page, page_size, include_profile,
        tags=_split_csv(tags), sort=sort, direction=dir,
        contains_col=contains_col, contains_val=contains_val,
    )


@router.post("/api/files/{file_id}/columns/fill", response_model=ReviewOut)
def fill_column(
    file_id: int,
    column: str,
    payload: ColumnFill,
    view: str = "all",
    tag: str | None = None,
    tags: str | None = None,
    sort: str | None = None,
    dir: str = "asc",
    contains_col: str | None = None,
    contains_val: str | None = None,
    page: int = 0,
    page_size: int = 50,
    include_profile: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Set a whole-column constant that fills every EMPTY cell of `column` with the
    given value — existing values are never overwritten. A blank value clears the
    constant. The column becomes an output column even if no input feeds it, so a
    batch-wide value (e.g. Revenue Share / Revenue Split) can be added in one go.
    Returns the refreshed Review payload."""
    f = _get_file_or_404(file_id, user, db)
    if column not in MASTER_COLUMN_TO_ATTR:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Unknown master column.")

    constants = dict(f.constants or {})
    value = payload.value.strip()
    if value:
        constants[column] = payload.value
    else:
        constants.pop(column, None)
    f.constants = constants
    db.commit()
    _invalidate(file_id)
    return _review_payload(
        f, view, tag, page, page_size, include_profile,
        tags=_split_csv(tags), sort=sort, direction=dir,
        contains_col=contains_col, contains_val=contains_val,
    )


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Sheet 1's name mirrors the Review tab the download came from.
_SHEET_NAMES = {
    "error": "Flagged rows",
    "auto_clean": "Auto cleaned",
    "manual_clean": "Manual cleaned",
    "clean": "Clean rows",
    "all": "Rows",
}

# Review-sheet cell highlights (Excel "Bad/Neutral" palette, filter-safe fills).
_FILL_EDITED = PatternFill("solid", fgColor="FFEB9C")  # yellow  -> human-corrected
_FILL_ERROR = PatternFill("solid", fgColor="FFC7CE")   # red     -> unresolved error
_HEADER_FONT = Font(bold=True)


def _cell_review(r: "CleanRow", manual_kind: str | None = None) -> tuple[dict[str, str], dict[str, str], str]:
    """Per-cell review annotations for one row.

    Returns (edited, errored, issue_text):
      - edited[col]   -> highlight the cell yellow; value is the resolved-issue note
        ("'old' -> 'new'"). A human changed it (inline edit / artist merge / bulk fill).
      - errored[col]  -> highlight the cell red; value is the error message.
      - issue_text    -> one-line, per-column summary of what was resolved / is still
        open, for the "Issue" column so a reviewer sees exactly what happened.

    `manual_kind` ("edited" / "kept") comes from the reviewer's actions, so a row
    kept as-is is called out even though none of its cells changed.
    """
    edited: dict[str, str] = {}
    errored: dict[str, str] = {}
    notes: list[str] = []
    for i in r.issues:
        col = i.get("column")
        if not col:
            continue
        if i["action"] == "error":
            msg = i.get("message") or i.get("tag") or "error"
            errored[col] = msg
            notes.append(f"{col}: {msg} (unresolved)")
        elif i.get("tag") == "corrected":
            old = i.get("original") or ""
            new = i.get("value") or ""
            edited[col] = f"'{old}' -> '{new}'"
            notes.append(f"{col}: '{old}' -> '{new}' (manually corrected)")
    if manual_kind == "kept" and not notes:
        notes.append("Kept as-is by a reviewer (values unchanged)")
    return edited, errored, "; ".join(notes)


def _write_review_sheet(
    ws,
    rows: list["CleanRow"],
    cols: list[str],
    manual: dict[int, str] | None = None,
) -> None:
    """Second workbook sheet: the same rows as the export, but with every cell a
    human touched highlighted yellow and every unresolved error highlighted red,
    plus an "Issue" column spelling out precisely what was resolved per cell.

    Streamed via a write-only worksheet (`ws` from a write_only Workbook) so even a
    six-figure-row file is produced in seconds with a few MB of memory — styled
    cells are `WriteOnlyCell`s built inline (random-access `ws.cell()` is not
    available in write-only mode)."""
    manual = manual or {}
    ws.freeze_panes = "A2"  # write-only: must be set before the first append
    header = []
    for value in ("Row", *cols, "Issue"):
        cell = WriteOnlyCell(ws, value=value)
        cell.font = _HEADER_FONT
        header.append(cell)
    header[-1].comment = Comment(
        "Yellow = manually corrected by a reviewer (e.g. artist renamed/merged).\n"
        "Red = error still unresolved.",
        "Cleanser",
    )
    ws.append(header)
    for r in rows:
        edited, errored, issue_text = _cell_review(r, manual.get(r.index))
        line = [r.index + 1]
        for col in cols:
            value = r.values.get(col, "")
            fill = _FILL_ERROR if col in errored else _FILL_EDITED if col in edited else None
            if fill is not None:
                cell = WriteOnlyCell(ws, value=value)
                cell.fill = fill
                line.append(cell)
            else:
                line.append(value)  # plain value -> no per-cell object overhead
        line.append(issue_text)
        ws.append(line)


@router.get("/api/files/{file_id}/export")
@limiter.limit(settings.heavy_rate_limit)
def export_rows(
    request: Request,
    file_id: int,
    view: str = "error",
    tag: str | None = None,
    tags: str | None = None,
    sort: str | None = None,
    dir: str = "asc",
    contains_col: str | None = None,
    contains_val: str | None = None,
    filename: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Download the currently-filtered rows as an .xlsx workbook.

    Honors the exact same view/tag/tags/sort/value-filter the Review grid is
    showing, so "download" gives the user the very set of rows in front of them
    (e.g. all rows sorted by Singer descending, or only the manually cleaned
    rows). Every tab exports the same two sheets: a plain one, and a review sheet
    where each cell a human changed is highlighted. The workbook is built in
    memory and streamed — nothing hits disk.
    """
    f = _get_file_or_404(file_id, user, db)
    manual = _manual_kinds(f)
    rows = _filter_rows(
        build_rows(f), view, tag, _split_csv(tags), contains_col, contains_val, manual
    )
    rows = _sort_rows(rows, sort, dir)
    cols = _active_columns(f)

    # write_only streams rows straight to the zip: a 15k-row export drops from
    # ~60s / ~130 MB (the default Workbook builds a styled object per cell, which
    # blew past the 60s proxy timeout — the reported "can't download") to ~3s / a
    # few MB, and scales to six-figure row counts well under any gateway timeout.
    wb = Workbook(write_only=True)
    # Sheet 1 — the plain export we already ship.
    ws = wb.create_sheet(_SHEET_NAMES.get(view, "Rows"))
    ws.append(["Row", *cols, "Issues"])
    for r in rows:
        issues = "; ".join(
            f"{i['column']}: {i.get('message') or i.get('tag')}"
            for i in r.issues
            if i.get("action") == "error"
        )
        ws.append([r.index + 1, *[r.values.get(c, "") for c in cols], issues])

    # Sheet 2 — same rows, but human-edited cells highlighted (yellow) and
    # unresolved errors (red), with an Issue column, so the export can be reviewed.
    _write_review_sheet(wb.create_sheet("Review (edits)"), rows, cols, manual)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Prefer the caller's descriptive name (built from the human-readable filter
    # labels on the client, e.g. "Demo(manually_edited_singers).xlsx"); otherwise
    # fall back to a self-explanatory default describing the view/sort/value.
    if filename:
        out_name = safe_filename(filename)
        if not out_name.lower().endswith(".xlsx"):
            out_name += ".xlsx"
    else:
        parts = ["flagged_rows" if view == "error" else f"{view}_rows"]
        if sort:
            parts.append(f"by_{sort}_{dir}")
        if contains_col and contains_val:
            parts.append(f"{contains_col}_{contains_val}")
        out_name = safe_filename("_".join(parts) + f"_file{file_id}.xlsx")
    return StreamingResponse(
        buf,
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": content_disposition(out_name)},
    )


@router.put("/api/files/{file_id}/rows/{row_index}", response_model=CleanRowOut)
def edit_row(
    file_id: int,
    row_index: int,
    payload: RowEdit,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Human edit: store the changed cells as a correction and re-clean the row."""
    f = _get_file_or_404(file_id, user, db)
    # Bounds-check against the warm cache (all row indices) when possible, so a
    # single edit doesn't force a reload of the whole data blob just to read len().
    base_values = _warm_base_values(f)
    if base_values is not None:
        if row_index not in base_values:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Row not found")
    elif not (0 <= row_index < len(f.data)):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Row not found")

    corrections = dict(f.corrections or {})
    existing = dict(corrections.get(str(row_index), {}))
    existing.update(payload.values)
    corrections[str(row_index)] = existing
    f.corrections = corrections
    db.commit()
    # Recompute just this row in place; fall back to a full rebuild if cache is cold.
    if not _edit_in_place(f, {row_index}):
        _invalidate(file_id)

    for r in build_rows(f):
        if r.index == row_index:
            return _row_out(r)
    raise HTTPException(status.HTTP_404_NOT_FOUND, "Row not found")


@router.put("/api/files/{file_id}/rows", response_model=ReviewOut)
def edit_rows(
    file_id: int,
    payload: RowsBatchEdit,
    view: str = "all",
    tag: str | None = None,
    tags: str | None = None,
    sort: str | None = None,
    dir: str = "asc",
    contains_col: str | None = None,
    contains_val: str | None = None,
    page: int = 0,
    page_size: int = 50,
    include_profile: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Apply many inline edits in one request and return the refreshed Review
    payload — so the grid updates in a single round-trip (no extra reload).

    Each edit is stored as a correction; rows that become fully clean afterwards
    drop out of the review queue automatically.
    """
    f = _get_file_or_404(file_id, user, db)
    corrections = dict(f.corrections or {})
    # Validate indices against the warm cache when possible (avoids loading the
    # data blob just to read its length); fall back to len(data) if cache is cold.
    base_values = _warm_base_values(f)
    n = len(base_values) if base_values is not None else len(f.data)
    changed: set[int] = set()
    for idx_str, values in payload.edits.items():
        try:
            idx = int(idx_str)
        except (TypeError, ValueError):
            continue
        if base_values is not None:
            if idx not in base_values:
                continue
        elif not (0 <= idx < n):
            continue
        existing = dict(corrections.get(idx_str, {}))
        existing.update(values)
        corrections[idx_str] = existing
        changed.add(idx)
    f.corrections = corrections
    db.commit()
    # Recompute only the edited rows in place; full rebuild only if cache is cold.
    if not _edit_in_place(f, changed):
        _invalidate(file_id)
    return _review_payload(
        f, view, tag, page, page_size, include_profile,
        tags=_split_csv(tags), sort=sort, direction=dir,
        contains_col=contains_col, contains_val=contains_val,
    )


@router.post("/api/files/{file_id}/accept", response_model=ReviewOut)
def accept_rows(
    file_id: int,
    payload: RowsAccept,
    view: str = "all",
    tag: str | None = None,
    tags: str | None = None,
    sort: str | None = None,
    dir: str = "asc",
    contains_col: str | None = None,
    contains_val: str | None = None,
    page: int = 0,
    page_size: int = 50,
    include_profile: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Keep flagged rows as-is: clear their error flags without changing values.

    `rows` is empty + `tag` given -> accept every row carrying that error type."""
    f = _get_file_or_404(file_id, user, db)
    accepted = set(f.accepted or [])
    if payload.rows:
        accepted.update(i for i in payload.rows if 0 <= i < len(f.data))
    elif tag:
        accepted.update(
            r.index for r in build_rows(f)
            if any(i.get("tag") == tag for i in r.issues)
        )
    f.accepted = sorted(accepted)
    db.commit()
    # Fast path: update the cached rows in place instead of a full re-clean.
    _accept_in_place(f)
    return _review_payload(
        f, view, tag, page, page_size, include_profile,
        tags=_split_csv(tags), sort=sort, direction=dir,
        contains_col=contains_col, contains_val=contains_val,
    )


@router.post("/api/files/{file_id}/revert", response_model=ReviewOut)
def revert_rows(
    file_id: int,
    payload: RowsRevert,
    view: str = "all",
    tag: str | None = None,
    tags: str | None = None,
    sort: str | None = None,
    dir: str = "asc",
    contains_col: str | None = None,
    contains_val: str | None = None,
    select_all: bool = False,
    page: int = 0,
    page_size: int = 50,
    include_profile: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Send manually-cleaned rows back to Needs review — the inverse of /accept.

    Drops each row's "keep as-is" acceptance AND its reviewer corrections, so the
    row re-cleans from its original values and its error flags come back. A row
    whose original values were already clean simply returns to the Auto cleaned
    tab. Rows the tool cleaned on its own are ignored (nothing to undo).

    `select_all=true` reverts every manually-cleaned row in the current filtered
    view instead of the explicit `rows` list."""
    f = _get_file_or_404(file_id, user, db)
    manual = _manual_kinds(f)
    if select_all:
        targets = {
            r.index
            for r in _filter_rows(
                build_rows(f), view, tag, _split_csv(tags),
                contains_col, contains_val, manual,
            )
            if r.index in manual
        }
    else:
        targets = {i for i in payload.rows if i in manual}

    if targets:
        f.accepted = [i for i in (f.accepted or []) if i not in targets]
        f.corrections = {
            k: v for k, v in (f.corrections or {}).items()
            if _row_key(k) not in targets
        }
        db.commit()
        # Re-clean only the reverted rows in the warm cache; the values change, so
        # this can't use the cheaper _accept_in_place path.
        if not _edit_in_place(f, targets):
            _invalidate(file_id)

    return _review_payload(
        f, view, tag, page, page_size, include_profile,
        tags=_split_csv(tags), sort=sort, direction=dir,
        contains_col=contains_col, contains_val=contains_val,
    )


@router.post("/api/files/{file_id}/clean/bulk", response_model=CleanSummary)
def bulk_fix(
    file_id: int,
    payload: BulkFix,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Resolve a whole class of errors at once (by tag).

    action="set"  -> set the flagged cells to `value` on every row carrying `tag`.
    action="drop" -> exclude every row carrying `tag` from the output.
    """
    f = _get_file_or_404(file_id, user, db)
    rows = build_rows(f)
    affected = [r for r in rows if any(i.get("tag") == payload.tag for i in r.issues)]

    if payload.action == "drop":
        dropped = set(f.dropped or [])
        dropped.update(r.index for r in affected)
        f.dropped = sorted(dropped)
    elif payload.action == "set":
        corrections = dict(f.corrections or {})
        for r in affected:
            override = dict(corrections.get(str(r.index), {}))
            if payload.column:
                override[payload.column] = payload.value or ""
            else:
                # Set every cell in this row that carries the tag — this correctly
                # handles tags that span multiple columns (e.g. two date columns).
                for i in r.issues:
                    if i.get("tag") == payload.tag:
                        override[i["column"]] = payload.value or ""
            corrections[str(r.index)] = override
        f.corrections = corrections
    else:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Unknown action.")

    db.commit()
    _invalidate(file_id)
    return _get_summary(f)


@router.post("/api/files/{file_id}/conflicts", response_model=ConflictsResult)
def check_conflicts(
    file_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Near-duplicates between this file's clean rows and the master dataset.

    A conflict is a clean row that all-but-matches an already-stored record
    (e.g. same singer / album / ISRC / UPC but a different composer or
    publisher). The frontend stacks each pair for cross-verification and asks the
    reviewer which is correct before the save proceeds. An empty list means the
    save can run straight through.
    """
    f = _get_file_or_404(file_id, user, db)
    clean = [r for r in build_rows(f) if r.status == "clean"]
    return ConflictsResult(
        conflicts=find_conflicts(db, clean),
        columns=list(MASTER_COLUMN_TO_ATTR),
    )


@router.post("/api/files/{file_id}/commit", response_model=CommitResult)
@limiter.limit(settings.heavy_rate_limit)
def commit_clean(
    request: Request,
    file_id: int,
    payload: CommitRequest | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Save all clean rows into the master dataset. Rows with errors are skipped.

    Cleaned rows land in the structured `master_data` table (one queryable
    column per master field), de-duplicated against everything already stored:
    an identical recording is not saved twice, and a row that matches except for
    its Label / Publisher / Distributor updates the existing record to the latest
    owner. Near-duplicates the reviewer flagged via `/conflicts` are applied per
    their decision (`resolutions`). Each save is recorded in the activity log.
    """
    f = _get_file_or_404(file_id, user, db)
    rows = build_rows(f)
    clean = [r for r in rows if r.status == "clean"]
    errors = len(rows) - len(clean)

    resolutions = {
        int(idx): {"decision": res.decision, "master_id": res.master_id}
        for idx, res in (payload.resolutions if payload else {}).items()
    }
    counts = upsert_master_records(db, f.branch_id, f.id, clean, resolutions)
    db.add(ActivityLog(
        branch_id=f.branch_id,
        file_id=f.id,
        action="commit",
        inserted=counts["inserted"],
        updated=counts["updated"],
        duplicates=counts["duplicates"],
        skipped_errors=errors,
    ))
    f.status = FileStatus.committed
    db.commit()
    return CommitResult(
        committed=counts["inserted"] + counts["updated"],
        skipped_errors=errors,
        inserted=counts["inserted"],
        updated=counts["updated"],
        duplicates=counts["duplicates"],
        skipped_conflicts=counts["skipped_conflicts"],
    )
